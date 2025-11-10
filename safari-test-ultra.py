import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry, Calendar
import sqlite3
import pandas as pd
from datetime import datetime, timedelta
import re
import shutil
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.cell.cell import MergedCell
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import hashlib
import secrets
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.backends import default_backend
import base64
import json


# --- Helpers: safe Excel column auto-adjust + footer & UI animation helpers ---
def auto_adjust_columns(ws, min_width=12, max_width=50):
    """Safely adjust column widths, skipping merged cells."""
    try:
        from openpyxl.cell.cell import MergedCell
        for col_cells in ws.columns:
            # find a representative non-merged cell
            first = None
            for cell in col_cells:
                if not isinstance(cell, MergedCell):
                    first = cell
                    break
            if first is None:
                continue
            column_letter = getattr(first, "column_letter", None)
            if not column_letter:
                continue
            max_length = 0
            for cell in col_cells:
                if isinstance(cell, MergedCell):
                    continue
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_length:
                    max_length = len(val)
            width = max(min_width, min(max_width, max_length + 2))
            try:
                ws.column_dimensions[column_letter].width = width
            except Exception:
                pass
    except Exception:
        pass

def _append_excel_footer(wb, sheet_name=None):
    """Append branded footer to workbook's active sheet or named sheet (merged across columns)."""
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        last_row = ws.max_row + 2
        max_col = ws.max_column or 1
        try:
            ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=max_col)
        except Exception:
            pass
        footer_cell = ws.cell(row=last_row, column=1)
        footer_cell.value = 'Developed by Mohammed Ismail'
        try:
            footer_cell.font = Font(italic=True)
            footer_cell.alignment = Alignment(horizontal='right')
        except Exception:
            pass
    except Exception:
        pass

def animate_glow(label, colors=None, delay=180):
    """Simple glow animation for CTkLabel (changes text_color periodically)."""
    try:
        if colors is None:
            colors = ['#B0BEC5', '#E0F7FF', '#FFFFFF', '#9BE7FF']
        idx = {'i': 0}
        def step():
            try:
                label.configure(text_color=colors[idx['i'] % len(colors)])
            except Exception:
                pass
            idx['i'] += 1
            try:
                label.after(delay, step)
            except Exception:
                pass
        label.after(delay, step)
    except Exception:
        pass
# --- end helpers ---

def _safe_set_col_width(ws, cell, width_value):
    try:
        col = getattr(cell, 'column_letter', None)
        if not col:
            return
        ws.column_dimensions[col].width = width_value
    except Exception:
        pass

def _append_excel_footer(wb, sheet_name=None):
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        max_row = ws.max_row + 2
        footer_cell = ws.cell(row=max_row, column=ws.max_column or 1)
        footer_cell.value = 'Developed by Mohammed Ismail'
        footer_cell.font = Font(italic=True)
        footer_cell.alignment = Alignment(horizontal='right')
    except Exception:
        pass


DB_PATH = 'wow_safari_booking.db'
BACKUP_DIR = 'backups'
USER = 'saurabh'
PASS = 'Sai@2001'
ADMIN_PIN = '1234'

# Encryption key for Excel files (store this securely in production)
EXCEL_PASSWORD = 'SafariBooking@2025'

# Email Configuration
EMAIL_CONFIG = {
    'smtp_server': 'smtp.gmail.com',
    'smtp_port': 587,
    'sender_email': 'your_email@gmail.com',
    'sender_password': 'your_app_password',
    'enabled': True  # Set to False if not configured
}

# SMS Configuration
SMS_CONFIG = {
    'api_url': 'https://api.smsgateway.com/send',
    'api_key': 'your_api_key',
    'enabled': False
}


class SecurityManager:
    """Enhanced security manager with encryption and hashing"""
    
    @staticmethod
    def hash_pin(pin):
        """Hash PIN for secure storage"""
        return hashlib.sha256(pin.encode()).hexdigest()
    
    @staticmethod
    def verify_pin(pin, hashed_pin):
        """Verify PIN against hash"""
        return SecurityManager.hash_pin(pin) == hashed_pin
    
    @staticmethod
    def generate_encryption_key(password):
        """Generate encryption key from password"""
        kdf = PBKDF2HMAC(
            algorithm=hashes.SHA256(),
            length=32,
            salt=b'safari_booking_salt',
            iterations=100000,
            backend=default_backend()
        )
        key = base64.urlsafe_b64encode(kdf.derive(password.encode()))
        return key
    
    @staticmethod
    def encrypt_data(data, password):
        """Encrypt sensitive data"""
        key = SecurityManager.generate_encryption_key(password)
        f = Fernet(key)
        return f.encrypt(data.encode()).decode()
    
    @staticmethod
    def decrypt_data(encrypted_data, password):
        """Decrypt sensitive data"""
        try:
            key = SecurityManager.generate_encryption_key(password)
            f = Fernet(key)
            return f.decrypt(encrypted_data.encode()).decode()
        except:
            return None


class Database:
    def __init__(self):
        self.conn = sqlite3.connect(DB_PATH, check_same_thread=False)
        self.create_tables()
        self.ensure_backup_dir()

    def ensure_backup_dir(self):
        if not os.path.exists(BACKUP_DIR):
            os.makedirs(BACKUP_DIR)

    def create_tables(self):
        cursor = self.conn.cursor()
        cursor.execute("PRAGMA table_info(bookings)")
        columns = [column[1] for column in cursor.fetchall()]
        
        self.conn.execute('''
        CREATE TABLE IF NOT EXISTS bookings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            members INTEGER NOT NULL,
            amount REAL NOT NULL,
            guest TEXT NOT NULL,
            phone TEXT NOT NULL,
            aadhaar TEXT NOT NULL,
            resort TEXT NOT NULL,
            interested TEXT DEFAULT 'No',
            activity TEXT DEFAULT 'None',
            profit REAL DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            notes TEXT,
            email TEXT,
            status TEXT DEFAULT 'Confirmed',
            payment_mode TEXT DEFAULT 'Cash',
            advance_paid REAL DEFAULT 0,
            balance_due REAL DEFAULT 0,
            check_in_time TEXT,
            check_out_time TEXT,
            special_requests TEXT
        )''')
        
        # Add missing columns
        new_columns = {
            'notes': 'TEXT',
            'email': 'TEXT',
            'status': "TEXT DEFAULT 'Confirmed'",
            'created_at': 'TEXT DEFAULT CURRENT_TIMESTAMP',
            'payment_mode': "TEXT DEFAULT 'Cash'",
            'advance_paid': 'REAL DEFAULT 0',
            'balance_due': 'REAL DEFAULT 0',
            'check_in_time': 'TEXT',
            'check_out_time': 'TEXT',
            'special_requests': 'TEXT'
        }
        
        for col_name, col_type in new_columns.items():
            if col_name not in columns:
                try:
                    self.conn.execute(f"ALTER TABLE bookings ADD COLUMN {col_name} {col_type}")
                except:
                    pass
        
        self.conn.execute('''
        CREATE TABLE IF NOT EXISTS activity_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action TEXT,
            details TEXT,
            user TEXT,
            timestamp TEXT DEFAULT CURRENT_TIMESTAMP
        )''')
        
        self.conn.execute('''
        CREATE TABLE IF NOT EXISTS security_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            event_type TEXT,
            details TEXT,
            ip_address TEXT,
            timestamp TEXT DEFAULT CURRENT_TIMESTAMP
        )''')
        
        self.conn.commit()

    def backup_database(self):
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_file = os.path.join(BACKUP_DIR, f'backup_{timestamp}.db')
            shutil.copy2(DB_PATH, backup_file)
            
            backups = sorted([f for f in os.listdir(BACKUP_DIR) if f.startswith('backup_')])
            if len(backups) > 30:
                for old_backup in backups[:-30]:
                    os.remove(os.path.join(BACKUP_DIR, old_backup))
            
            return backup_file
        except Exception as e:
            raise Exception(f"Backup failed: {str(e)}")

    def restore_database(self, backup_file):
        try:
            self.conn.close()
            shutil.copy2(backup_file, DB_PATH)
            self.conn = sqlite3.connect(DB_PATH, check_same_thread=False)
            return True
        except Exception as e:
            raise Exception(f"Restore failed: {str(e)}")

    def insert(self, booking):
        cur = self.conn.cursor()
        cur.execute('''INSERT INTO bookings 
            (date, members, amount, guest, phone, aadhaar, resort, interested, activity, 
             profit, notes, email, status, payment_mode, advance_paid, balance_due, 
             check_in_time, check_out_time, special_requests) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', booking)
        self.conn.commit()
        booking_id = cur.lastrowid
        self.log_activity("NEW_BOOKING", f"ID: {booking_id}, Guest: {booking[3]}, Resort: {booking[6]}, Amount: ₹{booking[2]}", USER)
        return booking_id

    def update_booking(self, booking_id, booking):
        cur = self.conn.cursor()
        cur.execute('''UPDATE bookings SET 
            date=?, members=?, amount=?, guest=?, phone=?, aadhaar=?, 
            resort=?, interested=?, activity=?, profit=?, notes=?, email=?, status=?,
            payment_mode=?, advance_paid=?, balance_due=?, check_in_time=?, 
            check_out_time=?, special_requests=?
            WHERE id=?''', booking + (booking_id,))
        self.conn.commit()
        self.log_activity("UPDATE_BOOKING", f"ID: {booking_id}, Guest: {booking[3]}", USER)

    def delete_booking(self, booking_id):
        cur = self.conn.cursor()
        cur.execute("SELECT guest, resort, amount FROM bookings WHERE id=?", (booking_id,))
        row = cur.fetchone()
        if row:
            cur.execute("DELETE FROM bookings WHERE id=?", (booking_id,))
            self.conn.commit()
            self.log_activity("DELETE_BOOKING", f"ID: {booking_id}, Guest: {row[0]}, Resort: {row[1]}, Amount: ₹{row[2]}", USER)
            return True
        return False

    def get_booking_by_id(self, booking_id):
        cur = self.conn.cursor()
        cur.execute("SELECT * FROM bookings WHERE id=?", (booking_id,))
        return cur.fetchone()

    def log_activity(self, action, details, user):
        cur = self.conn.cursor()
        cur.execute("INSERT INTO activity_log (action, details, user) VALUES (?, ?, ?)", 
                   (action, details, user))
        self.conn.commit()
    
    def log_security_event(self, event_type, details, ip_address="localhost"):
        cur = self.conn.cursor()
        cur.execute("INSERT INTO security_log (event_type, details, ip_address) VALUES (?, ?, ?)", 
                   (event_type, details, ip_address))
        self.conn.commit()

    def query_count(self, guest_filter="", resort_filter="", start_date=None, end_date=None, status_filter="", activity_filter=""):
        cur = self.conn.cursor()
        filters = []
        params = []
        
        if guest_filter:
            filters.append("(guest LIKE ? OR phone LIKE ?)")
            params.extend([f'%{guest_filter}%', f'%{guest_filter}%'])
        if resort_filter:
            filters.append("resort LIKE ?")
            params.append(f'%{resort_filter}%')
        if start_date:
            filters.append("date >= ?")
            params.append(start_date)
        if end_date:
            filters.append("date <= ?")
            params.append(end_date)
        if status_filter:
            filters.append("status = ?")
            params.append(status_filter)
        if activity_filter and activity_filter != "All":
            filters.append("activity = ?")
            params.append(activity_filter)

        query = "SELECT COUNT(*) FROM bookings"
        if filters:
            query += " WHERE " + " AND ".join(filters)
        
        cur.execute(query, params)
        return cur.fetchone()[0]

    def query_bookings(self, guest_filter="", resort_filter="", start_date=None, end_date=None, 
                      status_filter="", activity_filter="", limit=10, offset=0):
        cur = self.conn.cursor()
        filters = []
        params = []
        
        if guest_filter:
            filters.append("(guest LIKE ? OR phone LIKE ?)")
            params.extend([f'%{guest_filter}%', f'%{guest_filter}%'])
        if resort_filter:
            filters.append("resort LIKE ?")
            params.append(f'%{resort_filter}%')
        if start_date:
            filters.append("date >= ?")
            params.append(start_date)
        if end_date:
            filters.append("date <= ?")
            params.append(end_date)
        if status_filter:
            filters.append("status = ?")
            params.append(status_filter)
        if activity_filter and activity_filter != "All":
            filters.append("activity = ?")
            params.append(activity_filter)

        query = "SELECT * FROM bookings"
        if filters:
            query += " WHERE " + " AND ".join(filters)
        query += " ORDER BY date DESC, created_at DESC LIMIT ? OFFSET ?"
        params.extend([limit, offset])
        
        cur.execute(query, params)
        return cur.fetchall()

    def get_all_bookings(self, guest_filter="", resort_filter="", start_date=None, end_date=None, 
                        status_filter="", activity_filter=""):
        cur = self.conn.cursor()
        filters = []
        params = []
        
        if guest_filter:
            filters.append("(guest LIKE ? OR phone LIKE ?)")
            params.extend([f'%{guest_filter}%', f'%{guest_filter}%'])
        if resort_filter:
            filters.append("resort LIKE ?")
            params.append(f'%{resort_filter}%')
        if start_date:
            filters.append("date >= ?")
            params.append(start_date)
        if end_date:
            filters.append("date <= ?")
            params.append(end_date)
        if status_filter:
            filters.append("status = ?")
            params.append(status_filter)
        if activity_filter and activity_filter != "All":
            filters.append("activity = ?")
            params.append(activity_filter)

        query = "SELECT * FROM bookings"
        if filters:
            query += " WHERE " + " AND ".join(filters)
        query += " ORDER BY date DESC"
        
        cur.execute(query, params)
        return cur.fetchall()

    def summary(self, start_date=None, end_date=None):
        cur = self.conn.cursor()
        date_filter = ""
        params = []
        
        if start_date and end_date:
            date_filter = " WHERE date BETWEEN ? AND ?"
            params = [start_date, end_date]
        
        bookings = cur.execute(f"SELECT COUNT(*) FROM bookings{date_filter}", params).fetchone()[0] or 0
        revenue = cur.execute(f"SELECT SUM(amount) FROM bookings{date_filter}", params).fetchone()[0] or 0
        profit = cur.execute(f"SELECT SUM(profit) FROM bookings{date_filter}", params).fetchone()[0] or 0
        
        if date_filter:
            interested = cur.execute(f"SELECT COUNT(*) FROM bookings WHERE date BETWEEN ? AND ? AND interested='Yes'", params).fetchone()[0] or 0
        else:
            interested = cur.execute("SELECT COUNT(*) FROM bookings WHERE interested='Yes'").fetchone()[0] or 0
        
        advance = cur.execute(f"SELECT SUM(advance_paid) FROM bookings{date_filter}", params).fetchone()[0] or 0
        balance = cur.execute(f"SELECT SUM(balance_due) FROM bookings{date_filter}", params).fetchone()[0] or 0
        
        return {
            "bookings": bookings,
            "revenue": revenue,
            "profit": profit,
            "interested": interested,
            "advance_paid": advance,
            "balance_due": balance,
            "total_members": cur.execute(f"SELECT SUM(members) FROM bookings{date_filter}", params).fetchone()[0] or 0
        }
    
    def get_resort_stats(self):
        cur = self.conn.cursor()
        cur.execute("""
            SELECT resort, COUNT(*) as count, SUM(amount) as revenue, SUM(profit) as profit
            FROM bookings
            GROUP BY resort
            ORDER BY count DESC
        """)
        return cur.fetchall()

    def get_revenue_trend(self, days=30):
        cur = self.conn.cursor()
        cur.execute("""
            SELECT date, SUM(amount) as revenue, SUM(profit) as profit, COUNT(*) as bookings
            FROM bookings
            WHERE date >= date('now', '-' || ? || ' days')
            GROUP BY date
            ORDER BY date
        """, (days,))
        return cur.fetchall()

    def search_guest(self, phone):
        cur = self.conn.cursor()
        cur.execute("SELECT * FROM bookings WHERE phone = ? ORDER BY date DESC LIMIT 5", (phone,))
        return cur.fetchall()

    def get_activity_breakdown(self):
        cur = self.conn.cursor()
        cur.execute("""
            SELECT activity, COUNT(*) as count, SUM(amount) as revenue, SUM(profit) as profit
            FROM bookings
            WHERE activity != 'None' AND interested = 'Yes'
            GROUP BY activity
            ORDER BY count DESC
        """)
        return cur.fetchall()

    def close(self):
        self.conn.close()


class NotificationService:
    @staticmethod
    def send_email(to_email, subject, message, attachment_path=None):
        if not EMAIL_CONFIG['enabled'] or not to_email:
            return False
        
        try:
            msg = MIMEMultipart()
            msg['From'] = EMAIL_CONFIG['sender_email']
            msg['To'] = to_email
            msg['Subject'] = subject
            
            msg.attach(MIMEText(message, 'html'))
            
            if attachment_path and os.path.exists(attachment_path):
                with open(attachment_path, 'rb') as f:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(f.read())
                    encoders.encode_base64(part)
                    part.add_header('Content-Disposition', f'attachment; filename={os.path.basename(attachment_path)}')
                    msg.attach(part)
            
            server = smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port'])
            server.starttls()
            server.login(EMAIL_CONFIG['sender_email'], EMAIL_CONFIG['sender_password'])
            server.send_message(msg)
            server.quit()
            return True
        except Exception as e:
            print(f"Email error: {str(e)}")
            return False

    @staticmethod
    def send_booking_confirmation(booking_data):
        guest_name = booking_data['guest']
        resort = booking_data['resort']
        date = booking_data['date']
        members = booking_data['members']
        amount = booking_data['amount']
        booking_id = booking_data.get('booking_id', 'N/A')
        
        email_subject = f"🎉 Booking Confirmation #{booking_id} - Your Trip Our Place"
        email_body = f"""
        <html>
        <body style="font-family: 'Segoe UI', Arial, sans-serif; line-height: 1.8; color: #333; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 40px;">
            <div style="max-width: 650px; margin: 0 auto; background: white; border-radius: 20px; overflow: hidden; box-shadow: 0 10px 40px rgba(0,0,0,0.3);">
                <div style="background: linear-gradient(135deg, #00b4d8 0%, #0077b6 100%); padding: 40px 30px; text-align: center;">
                    <h1 style="color: white; margin: 0; font-size: 32px; text-shadow: 2px 2px 4px rgba(0,0,0,0.2);">🌴 Your Trip Our Place</h1>
                    <p style="color: rgba(255,255,255,0.95); margin: 10px 0 0 0; font-size: 16px;">Creating Unforgettable Memories</p>
                </div>
                
                <div style="padding: 40px 30px;">
                    <div style="text-align: center; margin-bottom: 30px;">
                        <div style="display: inline-block; background: #e8f5e9; color: #2e7d32; padding: 10px 25px; border-radius: 25px; font-weight: bold;">
                            ✅ BOOKING CONFIRMED
                        </div>
                    </div>
                    
                    <h2 style="color: #00b4d8; border-bottom: 3px solid #00b4d8; padding-bottom: 10px;">Hello {guest_name}! 👋</h2>
                    <p style="font-size: 16px; color: #555;">We're thrilled to confirm your booking. Get ready for an amazing experience!</p>
                    
                    <div style="background: linear-gradient(135deg, #e3f2fd 0%, #bbdefb 100%); padding: 25px; border-radius: 15px; margin: 25px 0; border-left: 5px solid #00b4d8;">
                        <h3 style="margin-top: 0; color: #0277bd;">📋 Booking Details</h3>
                        <table style="width: 100%; border-collapse: collapse;">
                            <tr>
                                <td style="padding: 10px 0; color: #666; font-weight: 600;">Booking ID:</td>
                                <td style="padding: 10px 0; color: #000; font-weight: bold;">#{booking_id}</td>
                            </tr>
                            <tr>
                                <td style="padding: 10px 0; color: #666; font-weight: 600;">🏨 Resort:</td>
                                <td style="padding: 10px 0; color: #000; font-weight: bold;">{resort}</td>
                            </tr>
                            <tr>
                                <td style="padding: 10px 0; color: #666; font-weight: 600;">📅 Date:</td>
                                <td style="padding: 10px 0; color: #000; font-weight: bold;">{date}</td>
                            </tr>
                            <tr>
                                <td style="padding: 10px 0; color: #666; font-weight: 600;">👥 Members:</td>
                                <td style="padding: 10px 0; color: #000; font-weight: bold;">{members}</td>
                            </tr>
                            <tr>
                                <td style="padding: 10px 0; color: #666; font-weight: 600;">💰 Amount:</td>
                                <td style="padding: 10px 0; color: #2e7d32; font-weight: bold; font-size: 18px;">₹{amount:,.2f}</td>
                            </tr>
                        </table>
                    </div>
                    
                    <div style="background: #fff3cd; border-left: 5px solid #ffc107; padding: 20px; border-radius: 10px; margin: 25px 0;">
                        <h4 style="margin-top: 0; color: #856404;">📌 Important Information</h4>
                        <ul style="margin: 10px 0; padding-left: 20px; color: #856404;">
                            <li>Please arrive 30 minutes before your scheduled time</li>
                            <li>Carry a valid ID proof for verification</li>
                            <li>Contact us for any changes or queries</li>
                        </ul>
                    </div>
                    
                    <div style="text-align: center; margin-top: 35px; padding-top: 25px; border-top: 2px dashed #ddd;">
                        <p style="color: #666; font-size: 14px; margin: 5px 0;">Need help? Contact us:</p>
                        <p style="color: #00b4d8; font-weight: bold; font-size: 16px; margin: 5px 0;">📞 +91-XXXXXXXXXX</p>
                        <p style="color: #00b4d8; font-weight: bold; font-size: 16px; margin: 5px 0;">📧 support@yourtripourplace.com</p>
                    </div>
                </div>
                
                <div style="background: #f8f9fa; padding: 25px 30px; text-align: center; border-top: 3px solid #00b4d8;">
                    <p style="margin: 0; color: #666; font-size: 14px;">Thank you for choosing Your Trip Our Place!</p>
                    <p style="margin: 10px 0 0 0; color: #999; font-size: 12px;">This is an automated email. Please do not reply.</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        return NotificationService.send_email(booking_data.get('email', ''), email_subject, email_body)


class ModernCalendarDialog(ctk.CTkToplevel):
    """Modern calendar picker dialog"""
    def __init__(self, parent, callback, initial_date=None):
        super().__init__(parent)
        self.callback = callback
        self.selected_date = initial_date or datetime.now().date()
        
        self.title("Select Date")
        self.geometry("400x450")
        self.resizable(False, False)
        self.configure(fg_color="#0a0e27")
        
        self.transient(parent)
        self.grab_set()
        
        # Header
        header = ctk.CTkFrame(self, fg_color="#1a2332", height=60)
        header.pack(fill="x", padx=20, pady=(20, 10))
        header.pack_propagate(False)
        
        ctk.CTkLabel(header, text="📅 Select Date", font=ctk.CTkFont(size=20, weight="bold"), 
                    text_color="#00d4ff").pack(pady=15)
        
        # Calendar
        cal_frame = ctk.CTkFrame(self, fg_color="#1a2332", corner_radius=15)
        cal_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        self.calendar = Calendar(cal_frame, selectmode='day', date_pattern='dd-mm-yyyy',
                                background='#1a2332', foreground='white',
                                selectbackground='#00b4d8', selectforeground='white',
                                normalbackground='#293241', normalforeground='white',
                                headersbackground='#00b4d8', headersforeground='white',
                                bordercolor='#00b4d8', borderwidth=2)
        
        if initial_date:
            self.calendar.selection_set(initial_date)
        
        self.calendar.pack(fill="both", expand=True, padx=15, pady=15)
        
        # Buttons
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(pady=15)
        
        ctk.CTkButton(btn_frame, text="✓ Select", command=self.select_date, 
                     width=120, height=40, fg_color="#00b4d8", hover_color="#0096c7",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(side="left", padx=5)
        ctk.CTkButton(btn_frame, text="✗ Cancel", command=self.destroy,
                     width=120, height=40, fg_color="#e63946", hover_color="#d62828",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(side="left", padx=5)
    
    def select_date(self):
        selected = self.calendar.get_date()
        self.callback(selected)
        self.destroy()


class PINDialog(ctk.CTkToplevel):
    def __init__(self, parent, callback):
        super().__init__(parent)
        self.callback = callback
        self.result = False
        
        self.title("Security Verification")
        self.geometry("450x350")
        self.resizable(False, False)
        self.configure(fg_color="#0a0e27")
        
        self.transient(parent)
        self.grab_set()
        
        container = ctk.CTkFrame(self, fg_color="#1a2332", corner_radius=20)
        container.pack(fill="both", expand=True, padx=30, pady=30)
        
        # Animated lock icon
        ctk.CTkLabel(container, text="🔐", font=ctk.CTkFont(size=56)).pack(pady=(25, 10))
        ctk.CTkLabel(container, text="Security PIN Required", 
                    font=ctk.CTkFont(size=22, weight="bold"), 
                    text_color="#00d4ff").pack(pady=(0, 8))
        ctk.CTkLabel(container, text="Enter PIN to access sensitive data", 
                    font=ctk.CTkFont(size=13), 
                    text_color="#98c1d9").pack(pady=(0, 25))
        
        # PIN entry with modern styling
        self.pin_entry = ctk.CTkEntry(container, placeholder_text="Enter 4-digit PIN", 
                                     show="●", width=220, height=45, 
                                     font=ctk.CTkFont(size=18), justify="center",
                                     border_color="#00b4d8", border_width=2)
        self.pin_entry.pack(pady=15)
        self.pin_entry.bind('<Return>', lambda e: self.verify_pin())
        self.pin_entry.focus()
        
        # Error label
        self.error_label = ctk.CTkLabel(container, text="", 
                                       font=ctk.CTkFont(size=11),
                                       text_color="#e63946")
        self.error_label.pack(pady=5)
        
        btn_frame = ctk.CTkFrame(container, fg_color="transparent")
        btn_frame.pack(pady=20)
        
        ctk.CTkButton(btn_frame, text="✓ Verify", command=self.verify_pin, 
                     width=130, height=42, fg_color="#00b4d8", hover_color="#0096c7",
                     font=ctk.CTkFont(size=15, weight="bold")).pack(side="left", padx=8)
        ctk.CTkButton(btn_frame, text="✗ Cancel", command=self.cancel, 
                     width=130, height=42, fg_color="#e63946", hover_color="#d62828",
                     font=ctk.CTkFont(size=15, weight="bold")).pack(side="left", padx=8)
        
        self.attempts = 0
        self.max_attempts = 3

    def verify_pin(self):
        entered_pin = self.pin_entry.get().strip()
        
        if entered_pin == ADMIN_PIN:
            self.result = True
            self.destroy()
            self.callback(self.result)
        else:
            self.attempts += 1
            remaining = self.max_attempts - self.attempts
            
            if remaining > 0:
                self.error_label.configure(text=f"❌ Incorrect PIN! {remaining} attempt(s) remaining")
                self.pin_entry.delete(0, 'end')
                self.pin_entry.configure(border_color="#e63946")
                self.after(100, lambda: self.pin_entry.configure(border_color="#00b4d8"))
                self.pin_entry.focus()
            else:
                messagebox.showerror("Access Denied", "Maximum attempts exceeded!\nReturning to main menu.", parent=self)
                self.cancel()

    def cancel(self):
        self.result = False
        self.destroy()
        self.callback(self.result)


class EditBookingDialog(ctk.CTkToplevel):
    def __init__(self, parent, booking_data, callback):
        super().__init__(parent)
        self.callback = callback
        self.booking_data = booking_data
        
        self.title("Edit Booking")
        self.geometry("900x800")
        self.configure(fg_color="#0a0e27")
        
        self.transient(parent)
        self.grab_set()
        
        # Header
        header = ctk.CTkFrame(self, fg_color="#1a2332", height=70)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        ctk.CTkLabel(header, text="✏️ Edit Booking", 
                    font=ctk.CTkFont(size=26, weight="bold"), 
                    text_color="#00d4ff").pack(pady=20)
        
        scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=30, pady=15)
        
        self.inputs = {}
        
        # Guest Information Section
        self.create_section(scroll, "👤 Guest Information")
        self.create_field(scroll, "Guest Name", booking_data[4])
        self.create_field(scroll, "Phone", booking_data[5])
        self.create_field(scroll, "Email", booking_data[13] if len(booking_data) > 13 else "")
        self.create_field(scroll, "Aadhaar", booking_data[6])
        
        # Booking Details Section
        self.create_section(scroll, "📅 Booking Details")
        self.create_field(scroll, "Resort", booking_data[7])
        self.create_field(scroll, "Date", booking_data[1])
        self.create_field(scroll, "Members", str(booking_data[2]))
        self.create_field(scroll, "Check-in Time", booking_data[18] if len(booking_data) > 18 else "")
        self.create_field(scroll, "Check-out Time", booking_data[19] if len(booking_data) > 19 else "")
        
        # Financial Details Section
        self.create_section(scroll, "💰 Financial Details")
        self.create_field(scroll, "Total Amount", str(booking_data[3]))
        self.create_field(scroll, "Advance Paid", str(booking_data[16] if len(booking_data) > 16 else 0))
        self.create_field(scroll, "Balance Due", str(booking_data[17] if len(booking_data) > 17 else 0))
        self.create_field(scroll, "Profit", str(booking_data[10]))
        
        # Payment Mode
        payment_frame = ctk.CTkFrame(scroll, fg_color="#293241", corner_radius=10)
        payment_frame.pack(fill="x", pady=5)
        ctk.CTkLabel(payment_frame, text="Payment Mode:", text_color="#98c1d9", 
                    font=ctk.CTkFont(weight="bold")).pack(anchor="w", padx=10, pady=(10, 5))
        self.payment_var = ctk.StringVar(value=booking_data[15] if len(booking_data) > 15 else "Cash")
        ctk.CTkComboBox(payment_frame, values=["Cash", "Card", "UPI", "Bank Transfer", "Cheque"], 
                       variable=self.payment_var, width=200).pack(anchor="w", padx=10, pady=(0, 10))
        
        # Status
        status_frame = ctk.CTkFrame(scroll, fg_color="#293241", corner_radius=10)
        status_frame.pack(fill="x", pady=5)
        ctk.CTkLabel(status_frame, text="Status:", text_color="#98c1d9", 
                    font=ctk.CTkFont(weight="bold")).pack(anchor="w", padx=10, pady=(10, 5))
        self.status_var = ctk.StringVar(value=booking_data[14] if len(booking_data) > 14 else "Confirmed")
        ctk.CTkComboBox(status_frame, values=["Confirmed", "Pending", "Cancelled", "Completed"], 
                       variable=self.status_var, width=200).pack(anchor="w", padx=10, pady=(0, 10))
        
        # Activity Section
        self.create_section(scroll, "🌊 Water Activities")
        activity_frame = ctk.CTkFrame(scroll, fg_color="#293241", corner_radius=10)
        activity_frame.pack(fill="x", pady=5)
        ctk.CTkLabel(activity_frame, text="Water Activity:", text_color="#98c1d9", 
                    font=ctk.CTkFont(weight="bold")).pack(anchor="w", padx=10, pady=(10, 5))
        self.activity_var = ctk.StringVar(value=booking_data[9])
        ctk.CTkComboBox(activity_frame, 
                       values=["None", "Kayaking", "Zorbing", "Swimming", "Boating", 
                              "Zip Line", "Short Rafting", "Mid Rafting", "Long Rafting", 
                              "Go Karting", "All Activities"], 
                       variable=self.activity_var, width=250).pack(anchor="w", padx=10, pady=(0, 10))
        
        # Notes Section
        self.create_section(scroll, "📝 Additional Information")
        notes_frame = ctk.CTkFrame(scroll, fg_color="#293241", corner_radius=10)
        notes_frame.pack(fill="x", pady=5)
        ctk.CTkLabel(notes_frame, text="Notes:", text_color="#98c1d9", 
                    font=ctk.CTkFont(weight="bold")).pack(anchor="w", padx=10, pady=(10, 5))
        self.notes_text = ctk.CTkTextbox(notes_frame, height=80, font=ctk.CTkFont(size=12))
        self.notes_text.pack(fill="x", padx=10, pady=(0, 10))
        if len(booking_data) > 12 and booking_data[12]:
            self.notes_text.insert("1.0", booking_data[12])
        
        # Special Requests
        ctk.CTkLabel(notes_frame, text="Special Requests:", text_color="#98c1d9", 
                    font=ctk.CTkFont(weight="bold")).pack(anchor="w", padx=10, pady=(10, 5))
        self.special_requests_text = ctk.CTkTextbox(notes_frame, height=80, font=ctk.CTkFont(size=12))
        self.special_requests_text.pack(fill="x", padx=10, pady=(0, 10))
        if len(booking_data) > 20 and booking_data[20]:
            self.special_requests_text.insert("1.0", booking_data[20])
        
        # Buttons
        btn_frame = ctk.CTkFrame(scroll, fg_color="transparent")
        btn_frame.pack(pady=25)
        
        ctk.CTkButton(btn_frame, text="💾 Save Changes", command=self.save_changes, 
                     width=160, height=45, fg_color="#00b4d8", hover_color="#0096c7",
                     font=ctk.CTkFont(size=15, weight="bold")).pack(side="left", padx=10)
        ctk.CTkButton(btn_frame, text="✗ Cancel", command=self.destroy,
                     width=160, height=45, fg_color="#e63946", hover_color="#d62828",
                     font=ctk.CTkFont(size=15, weight="bold")).pack(side="left", padx=10)

    def create_section(self, parent, title):
        section_label = ctk.CTkLabel(parent, text=title, 
                                     font=ctk.CTkFont(size=18, weight="bold"), 
                                     text_color="#00d4ff")
        section_label.pack(anchor="w", padx=10, pady=(20, 10))

    def create_field(self, parent, label, value):
        frame = ctk.CTkFrame(parent, fg_color="#293241", corner_radius=10)
        frame.pack(fill="x", pady=5)
        ctk.CTkLabel(frame, text=f"{label}:", text_color="#98c1d9", 
                    font=ctk.CTkFont(weight="bold")).pack(anchor="w", padx=10, pady=(10, 5))
        entry = ctk.CTkEntry(frame, width=450, height=38, font=ctk.CTkFont(size=13))
        entry.insert(0, str(value))
        entry.pack(anchor="w", padx=10, pady=(0, 10))
        self.inputs[label] = entry

    def save_changes(self):
        try:
            members = int(self.inputs["Members"].get())
            amount = float(self.inputs["Total Amount"].get())
            profit = float(self.inputs["Profit"].get())
            advance = float(self.inputs["Advance Paid"].get())
            balance = float(self.inputs["Balance Due"].get())
            
            if members <= 0 or amount < 0 or profit < 0:
                messagebox.showerror("Invalid Input", "Please check your numeric values!", parent=self)
                return
            
            updated_data = {
                'date': self.inputs["Date"].get(),
                'members': members,
                'amount': amount,
                'guest': self.inputs["Guest Name"].get(),
                'phone': self.inputs["Phone"].get(),
                'aadhaar': self.inputs["Aadhaar"].get(),
                'resort': self.inputs["Resort"].get(),
                'interested': 'Yes' if self.activity_var.get() != 'None' else 'No',
                'activity': self.activity_var.get(),
                'profit': profit,
                'notes': self.notes_text.get("1.0", "end-1c"),
                'email': self.inputs["Email"].get(),
                'status': self.status_var.get(),
                'payment_mode': self.payment_var.get(),
                'advance_paid': advance,
                'balance_due': balance,
                'check_in_time': self.inputs["Check-in Time"].get(),
                'check_out_time': self.inputs["Check-out Time"].get(),
                'special_requests': self.special_requests_text.get("1.0", "end-1c")
            }
            
            self.callback(updated_data)
            self.destroy()
            
        except ValueError:
            messagebox.showerror("Invalid Input", "Please enter valid numbers for numeric fields!", parent=self)


class LoginPage(ctk.CTkFrame):
    def __init__(self, master, app):
        super().__init__(master)
        self.app = app
        self.configure(fg_color="#0a0e27")
        
        container = ctk.CTkFrame(self, fg_color="transparent")
        container.place(relx=0.5, rely=0.5, anchor="center")

        glass_card = ctk.CTkFrame(container, fg_color="#1a2332", corner_radius=25, 
                                 border_width=2, border_color="#3d5a80")
        glass_card.pack(padx=60, pady=40)

        ctk.CTkLabel(glass_card, text="🌴 Your Trip Our Place", 
                    font=ctk.CTkFont(size=34, weight="bold"), 
                    text_color="#00d4ff").pack(pady=(45, 10))
        ctk.CTkLabel(glass_card, text="Professional Edition v2.0", 
                    font=ctk.CTkFont(size=15), 
                    text_color="#98c1d9").pack(pady=(0, 25))

        form_frame = ctk.CTkFrame(glass_card, fg_color="transparent")
        form_frame.pack(pady=(25, 0), padx=55)

        ctk.CTkLabel(form_frame, text="Welcome Back", 
                    font=ctk.CTkFont(size=22, weight="bold"), 
                    text_color="#ffffff").pack(pady=(0, 25))

        ctk.CTkLabel(form_frame, text="Username", 
                    text_color="#98c1d9", 
                    font=ctk.CTkFont(size=13)).pack(anchor="w", pady=(0, 6))
        self.username_input = ctk.CTkEntry(form_frame, placeholder_text="Enter username", 
                                          width=320, height=42, 
                                          font=ctk.CTkFont(size=14))
        self.username_input.pack(pady=(0, 18))

        ctk.CTkLabel(form_frame, text="Password", 
                    text_color="#98c1d9", 
                    font=ctk.CTkFont(size=13)).pack(anchor="w", pady=(0, 6))
        self.password_input = ctk.CTkEntry(form_frame, placeholder_text="Enter password", 
                                          show="●", width=320, height=42,
                                          font=ctk.CTkFont(size=14))
        self.password_input.pack(pady=(0, 25))
        self.password_input.bind('<Return>', lambda e: self.check_login())

        ctk.CTkButton(form_frame, text="🔐 Login", command=self.check_login, 
                     width=320, height=48, font=ctk.CTkFont(size=17, weight="bold"), 
                     fg_color="#00b4d8", hover_color="#0096c7").pack(pady=(0, 45))

    def check_login(self):
        u = self.username_input.get().strip()
        p = self.password_input.get().strip()
        if u == USER and p == PASS:
            self.app.show_main_app()
        else:
            messagebox.showerror("Login Failed", "Invalid credentials!")
            self.password_input.delete(0, 'end')


class BookingApp(ctk.CTk):
    def add_developer_footer(self, parent):
        try:
            footer = ctk.CTkFrame(parent, fg_color='transparent', height=28)
            footer.pack(side='bottom', fill='x')
            lbl = ctk.CTkLabel(footer, text='Developed by Mohammed Ismail', font=ctk.CTkFont(size=14, slant='italic'), text_color='#B0BEC5')
            lbl.pack(pady=6)
            try:
                animate_glow(lbl, colors=['#6dd3ff','#9be7ff','#d1f5ff'], delay=200)
            except Exception:
                pass
        except Exception:
            pass

    def __init__(self):
        super().__init__()
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("dark-blue")

        self.title("Safari Booking System - Professional Edition v2.0 — Developed by Mohammed Ismail")
        self.geometry("1450x920")
        self.minsize(1300, 800)

        self.db = Database()
        self.current_page = 1
        self.page_size = 20
        self.configure(fg_color="#0a0e27")
        self.authenticated_pages = set()

        self.RESORT_LIST = [
            "Hidden valley", "AARKAY RESORT", "MANOMAYA", "LAKE VALLEY", "Power Star",
            "RAJHANS RESORT", "CRYSTAL GREEN JUNGLE STAY", "STARLIGHT RESORT River Ret",
            "9 Coins Resort", "Century Resort", "Golden Egle", "white Elephant",
            "Fire flies", "WILD inn jungle Home Stay", "Rao Saheb Villa", "King's Resort",
            "Greenland Jungle Stay", "Fly Catcher", "infinity resort", "Wildernest Jungle Resort",
            "CROCODILE EDGE HOME STAY", "River Point Resort", "Woodpecker resort",
            "PEPPER PARADISE HOMESTAY", "Whistling woods", "Boutique Resort",
            "Cashew Jungle Stay", "Jungle Resorts", "Tiger", "Panther Resorts",
            "--- Enter Custom Resort ---"
        ]

        self.main_container = ctk.CTkFrame(self, fg_color="#0a0e27")
        self.main_container.pack(fill="both", expand=True)

        self.login_page = LoginPage(self.main_container, self)
        self.login_page.pack(fill="both", expand=True)

        self.attributes('-fullscreen', True)
        self.bind('<Escape>', self.toggle_fullscreen)
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

    def show_main_app(self):
        self.login_page.pack_forget()
        self.login_page.destroy()
        
        self.create_navigation()
        
        self.content_area = ctk.CTkFrame(self.main_container, fg_color="#0a0e27")
        self.content_area.pack(fill="both", expand=True)
        try:
            self.add_developer_footer(self.content_area)
        except Exception:
            pass
        
        self.booking_frame = ctk.CTkFrame(self.content_area, fg_color="transparent")
        self.history_frame = ctk.CTkFrame(self.content_area, fg_color="transparent")
        self.summary_frame = ctk.CTkFrame(self.content_area, fg_color="transparent")
        self.backup_frame = ctk.CTkFrame(self.content_area, fg_color="transparent")
        
        self.booking_frame.pack(fill="both", expand=True)
        self.history_frame.pack(fill="both", expand=True)
        self.summary_frame.pack(fill="both", expand=True)
        self.backup_frame.pack(fill="both", expand=True)
        
        self.create_booking_ui()
        self.create_history_ui()
        self.create_summary_ui()
        self.create_backup_ui()
        
        self.show_booking()
        self.auto_backup()

    def auto_backup(self):
        try:
            self.db.backup_database()
        except:
            pass

    def create_navigation(self):
        nav_frame = ctk.CTkFrame(self.main_container, height=95, fg_color="#1a2332")
        nav_frame.pack(side="top", fill="x")
        nav_frame.pack_propagate(False)

        ctk.CTkLabel(nav_frame, text="🌴 Your Trip Our Place", 
                    font=ctk.CTkFont(size=30, weight="bold"), 
                    text_color="#00d4ff").pack(side="left", padx=35, pady=22)

        btn_container = ctk.CTkFrame(nav_frame, fg_color="transparent")
        btn_container.pack(side="left", padx=60)

        buttons = [
            ("📝 New Booking", self.show_booking, "#00b4d8"),
            ("📚 History", self.show_history_secure, "#0077b6"),
            ("📊 Analytics", self.show_summary_secure, "#023e8a"),
            ("💾 Backup", self.show_backup, "#6a4c93")
        ]

        for text, cmd, color in buttons:
            ctk.CTkButton(btn_container, text=text, command=cmd, 
                         width=155, height=42, fg_color=color, 
                         font=ctk.CTkFont(size=15, weight="bold"),
                         corner_radius=10).pack(side="left", padx=6)

        ctk.CTkButton(nav_frame, text="🚪 Logout", command=self.logout, 
                     width=110, height=42, fg_color="#e63946", hover_color="#d62828",
                     font=ctk.CTkFont(size=15, weight="bold"),
                     corner_radius=10).pack(side="right", padx=35)

        try:
            self.add_developer_footer(nav_frame)
        except Exception:
            pass

    def logout(self):
        if messagebox.askyesno("Logout", "Are you sure you want to logout?"):
            self.authenticated_pages.clear()
            self.main_container.destroy()
            self.main_container = ctk.CTkFrame(self, fg_color="#0a0e27")
            self.main_container.pack(fill="both", expand=True)
            self.login_page = LoginPage(self.main_container, self)
            self.login_page.pack(fill="both", expand=True)

    def toggle_fullscreen(self, event=None):
        current = self.attributes('-fullscreen')
        self.attributes('-fullscreen', not current)

    def show_history_secure(self):
        if 'history' not in self.authenticated_pages:
            PINDialog(self, lambda success: self.handle_auth('history', success) if success else None)
        else:
            self.show_history()

    def show_summary_secure(self):
        if 'analytics' not in self.authenticated_pages:
            PINDialog(self, lambda success: self.handle_auth('analytics', success) if success else None)
        else:
            self.show_summary()

    def handle_auth(self, page, success):
        if success:
            self.authenticated_pages.add(page)
            self.db.log_security_event(f"ACCESS_GRANTED_{page.upper()}", f"User accessed {page} section")
            if page == 'history':
                self.show_history()
            elif page == 'analytics':
                self.show_summary()

    def show_booking(self):
        self.history_frame.pack_forget()
        self.summary_frame.pack_forget()
        self.backup_frame.pack_forget()
        self.booking_frame.pack(fill="both", expand=True)

    def show_history(self):
        self.booking_frame.pack_forget()
        self.summary_frame.pack_forget()
        self.backup_frame.pack_forget()
        self.history_frame.pack(fill="both", expand=True)
        self.current_page = 1
        self.load_history()

    def show_summary(self):
        self.booking_frame.pack_forget()
        self.history_frame.pack_forget()
        self.backup_frame.pack_forget()
        self.summary_frame.pack(fill="both", expand=True)
        self.load_summary()

    def show_backup(self):
        self.booking_frame.pack_forget()
        self.history_frame.pack_forget()
        self.summary_frame.pack_forget()
        self.backup_frame.pack(fill="both", expand=True)
        self.load_backup_list()

    def create_booking_ui(self):
        for widget in self.booking_frame.winfo_children():
            widget.destroy()

        # Header with gradient effect
        header = ctk.CTkFrame(self.booking_frame, fg_color="#1a2332", height=80)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        ctk.CTkLabel(header, text="✨ Create New Booking", 
                    font=ctk.CTkFont(size=30, weight="bold"), 
                    text_color="#00d4ff").pack(pady=25)

        scroll = ctk.CTkScrollableFrame(self.booking_frame, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=45, pady=15)

        card = ctk.CTkFrame(scroll, fg_color="#1a2332", corner_radius=18)
        card.pack(fill="both", padx=25, pady=15)

        form = ctk.CTkFrame(card, fg_color="transparent")
        form.pack(padx=35, pady=35, fill="both")

        self.booking_inputs = {}
        
        # Guest Information
        self.create_booking_section(form, "👤 Guest Information")
        guest_container = ctk.CTkFrame(form, fg_color="transparent")
        guest_container.pack(fill="x", pady=(0, 20))
        
        # Two-column layout
        left_col = ctk.CTkFrame(guest_container, fg_color="transparent")
        left_col.pack(side="left", fill="both", expand=True, padx=(0, 15))
        
        right_col = ctk.CTkFrame(guest_container, fg_color="transparent")
        right_col.pack(side="left", fill="both", expand=True, padx=(15, 0))
        
        self.create_modern_field(left_col, "* Phone Number", "Phone Number", 
                                "Enter 10-digit phone number", bind_func=self.check_phone_length)
        self.create_modern_field(left_col, "* Guest Name", "Guest Name", "Enter full name")
        
        self.create_modern_field(right_col, "Email Address", "Email", "guest@example.com")
        self.create_modern_field(right_col, "* Aadhaar Number", "Aadhaar Card Number", 
                                "Enter 12-digit Aadhaar number")

        # Booking Details
        self.create_booking_section(form, "📅 Booking Details")
        booking_container = ctk.CTkFrame(form, fg_color="transparent")
        booking_container.pack(fill="x", pady=(0, 20))
        
        booking_left = ctk.CTkFrame(booking_container, fg_color="transparent")
        booking_left.pack(side="left", fill="both", expand=True, padx=(0, 15))
        
        booking_right = ctk.CTkFrame(booking_container, fg_color="transparent")
        booking_right.pack(side="left", fill="both", expand=True, padx=(15, 0))
        
        # Modern Date Picker
        date_frame = ctk.CTkFrame(booking_left, fg_color="#293241", corner_radius=12)
        date_frame.pack(fill="x", pady=8)
        ctk.CTkLabel(date_frame, text="* Booking Date", text_color="#98c1d9", 
                    font=ctk.CTkFont(weight="bold", size=13)).pack(anchor="w", padx=15, pady=(12, 6))
        
        date_picker_frame = ctk.CTkFrame(date_frame, fg_color="transparent")
        date_picker_frame.pack(fill="x", padx=15, pady=(0, 12))
        
        self.date_display = ctk.CTkEntry(date_picker_frame, height=40, 
                                        font=ctk.CTkFont(size=13),
                                        state="readonly")
        self.date_display.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.date_display.configure(state="normal")
        self.date_display.insert(0, datetime.now().strftime("%d-%m-%Y"))
        self.date_display.configure(state="readonly")
        
        self.selected_date = datetime.now().date()
        
        ctk.CTkButton(date_picker_frame, text="📅", command=self.open_calendar_picker,
                     width=45, height=40, fg_color="#00b4d8", hover_color="#0096c7",
                     font=ctk.CTkFont(size=18)).pack(side="left")
        
        # Resort Selection
        resort_frame = ctk.CTkFrame(booking_left, fg_color="#293241", corner_radius=12)
        resort_frame.pack(fill="x", pady=8)
        ctk.CTkLabel(resort_frame, text="* Resort", text_color="#98c1d9", 
                    font=ctk.CTkFont(weight="bold", size=13)).pack(anchor="w", padx=15, pady=(12, 6))
        
        self.resort_var = ctk.StringVar(value=self.RESORT_LIST[0])
        self.resort_combo = ctk.CTkComboBox(resort_frame, values=self.RESORT_LIST, 
                                           variable=self.resort_var, width=300, height=40,
                                           font=ctk.CTkFont(size=13),
                                           command=self.on_resort_change)
        self.resort_combo.pack(anchor="w", padx=15, pady=(0, 8))
        
        self.custom_resort_entry = ctk.CTkEntry(resort_frame, placeholder_text="Enter custom resort name...", 
                                               width=300, height=40, state="disabled",
                                               font=ctk.CTkFont(size=13))
        self.custom_resort_entry.pack(anchor="w", padx=15, pady=(0, 12))
        
        # Check-in/out times
        self.create_modern_field(booking_right, "Check-in Time", "Check-in Time", "e.g., 02:00 PM")
        self.create_modern_field(booking_right, "Check-out Time", "Check-out Time", "e.g., 11:00 AM")

        # Financial Details
        self.create_booking_section(form, "💰 Financial Details")
        financial_container = ctk.CTkFrame(form, fg_color="transparent")
        financial_container.pack(fill="x", pady=(0, 20))
        
        fin_left = ctk.CTkFrame(financial_container, fg_color="transparent")
        fin_left.pack(side="left", fill="both", expand=True, padx=(0, 15))
        
        fin_right = ctk.CTkFrame(financial_container, fg_color="transparent")
        fin_right.pack(side="left", fill="both", expand=True, padx=(15, 0))
        
        self.create_modern_field(fin_left, "* Number of Members", "Members", "Must be greater than 0")
        self.create_modern_field(fin_left, "* Total Amount (₹)", "Amount Paid", "Enter amount in rupees")
        
        self.create_modern_field(fin_right, "* Advance Paid (₹)", "Advance Paid", "Enter advance amount")
        self.create_modern_field(fin_right, "* Balance Due (₹)", "Balance Due", "Enter balance due")
        self.create_modern_field(fin_right, "* Profit (₹)", "Profit", "Enter profit in rupees")
        
        # Payment Mode
        payment_frame = ctk.CTkFrame(fin_left, fg_color="#293241", corner_radius=12)
        payment_frame.pack(fill="x", pady=8)
        ctk.CTkLabel(payment_frame, text="Payment Mode", text_color="#98c1d9", 
                    font=ctk.CTkFont(weight="bold", size=13)).pack(anchor="w", padx=15, pady=(12, 6))
        self.payment_var = ctk.StringVar(value="Cash")
        ctk.CTkComboBox(payment_frame, values=["Cash", "Card", "UPI", "Bank Transfer", "Cheque"], 
                       variable=self.payment_var, width=250, height=40,
                       font=ctk.CTkFont(size=13)).pack(anchor="w", padx=15, pady=(0, 12))

        # Activities
        self.create_booking_section(form, "🌊 Water Activities")
        activity_container = ctk.CTkFrame(form, fg_color="transparent")
        activity_container.pack(fill="x", pady=(0, 20))
        
        act_left = ctk.CTkFrame(activity_container, fg_color="transparent")
        act_left.pack(side="left", fill="both", expand=True, padx=(0, 15))
        
        act_right = ctk.CTkFrame(activity_container, fg_color="transparent")
        act_right.pack(side="left", fill="both", expand=True, padx=(15, 0))
        
        interest_frame = ctk.CTkFrame(act_left, fg_color="#293241", corner_radius=12)
        interest_frame.pack(fill="x", pady=8)
        ctk.CTkLabel(interest_frame, text="Interested in Water Activities?", text_color="#98c1d9", 
                    font=ctk.CTkFont(weight="bold", size=13)).pack(anchor="w", padx=15, pady=(12, 6))
        self.interested_var = ctk.StringVar(value="No")
        ctk.CTkComboBox(interest_frame, values=["No", "Yes"], variable=self.interested_var, 
                       width=140, height=40, font=ctk.CTkFont(size=13),
                       command=self.toggle_activity_dropdown).pack(anchor="w", padx=15, pady=(0, 12))

        activity_frame = ctk.CTkFrame(act_right, fg_color="#293241", corner_radius=12)
        activity_frame.pack(fill="x", pady=8)
        ctk.CTkLabel(activity_frame, text="Select Activity:", text_color="#98c1d9", 
                    font=ctk.CTkFont(weight="bold", size=13)).pack(anchor="w", padx=15, pady=(12, 6))
        self.activity_var = ctk.StringVar(value="None")
        self.activity_combo = ctk.CTkComboBox(activity_frame, 
                       values=["None", "Kayaking", "Zorbing", "Swimming", "Boating", 
                              "Zip Line", "Short Rafting", "Mid Rafting", "Long Rafting", 
                              "Go Karting", "All Activities"], 
                       variable=self.activity_var, width=280, height=40, state="disabled",
                       font=ctk.CTkFont(size=13))
        self.activity_combo.pack(anchor="w", padx=15, pady=(0, 12))

        # Status & Notes
        self.create_booking_section(form, "📝 Additional Information")
        notes_container = ctk.CTkFrame(form, fg_color="#293241", corner_radius=12)
        notes_container.pack(fill="x", pady=(0, 20))
        
        status_frame = ctk.CTkFrame(notes_container, fg_color="transparent")
        status_frame.pack(fill="x", padx=15, pady=(15, 10))
        
        ctk.CTkLabel(status_frame, text="Status:", text_color="#98c1d9", 
                    font=ctk.CTkFont(weight="bold", size=13)).pack(anchor="w", pady=(0, 6))
        self.status_var = ctk.StringVar(value="Confirmed")
        ctk.CTkComboBox(status_frame, values=["Confirmed", "Pending", "Cancelled", "Completed"], 
                       variable=self.status_var, width=220, height=40,
                       font=ctk.CTkFont(size=13)).pack(anchor="w")

        ctk.CTkLabel(notes_container, text="Notes:", text_color="#98c1d9", 
                    font=ctk.CTkFont(weight="bold", size=13)).pack(anchor="w", padx=15, pady=(15, 6))
        self.notes_textbox = ctk.CTkTextbox(notes_container, height=90, font=ctk.CTkFont(size=13))
        self.notes_textbox.pack(fill="x", padx=15, pady=(0, 10))
        
        ctk.CTkLabel(notes_container, text="Special Requests:", text_color="#98c1d9", 
                    font=ctk.CTkFont(weight="bold", size=13)).pack(anchor="w", padx=15, pady=(10, 6))
        self.special_requests_textbox = ctk.CTkTextbox(notes_container, height=90, 
                                                       font=ctk.CTkFont(size=13))
        self.special_requests_textbox.pack(fill="x", padx=15, pady=(0, 15))

        # Notification Options
        notification_frame = ctk.CTkFrame(form, fg_color="#293241", corner_radius=12)
        notification_frame.pack(fill="x", pady=(0, 25))
        
        ctk.CTkLabel(notification_frame, text="📬 Notification Preferences", 
                    text_color="#00d4ff", 
                    font=ctk.CTkFont(size=16, weight="bold")).pack(anchor="w", padx=15, pady=(15, 10))
        
        self.send_email_var = ctk.BooleanVar(value=EMAIL_CONFIG['enabled'])
        self.send_sms_var = ctk.BooleanVar(value=SMS_CONFIG['enabled'])
        
        ctk.CTkCheckBox(notification_frame, text="📧 Send Email Confirmation", 
                       variable=self.send_email_var, text_color="#98c1d9",
                       font=ctk.CTkFont(size=13),
                       checkbox_width=22, checkbox_height=22).pack(anchor="w", padx=15, pady=5)
        ctk.CTkCheckBox(notification_frame, text="📱 Send SMS Confirmation", 
                       variable=self.send_sms_var, text_color="#98c1d9",
                       font=ctk.CTkFont(size=13),
                       checkbox_width=22, checkbox_height=22).pack(anchor="w", padx=15, pady=(5, 15))

        # Save Button
        ctk.CTkButton(form, text="💾 Save Booking", command=self.save_booking, 
                     width=320, height=55, font=ctk.CTkFont(size=18, weight="bold"), 
                     fg_color="#00b4d8", hover_color="#0096c7",
                     corner_radius=12).pack(pady=25)

    def create_booking_section(self, parent, title):
        section_frame = ctk.CTkFrame(parent, fg_color="transparent")
        section_frame.pack(fill="x", pady=(20, 12))
        
        ctk.CTkLabel(section_frame, text=title, 
                    font=ctk.CTkFont(size=19, weight="bold"), 
                    text_color="#00d4ff").pack(anchor="w", padx=5)
        
        separator = ctk.CTkFrame(section_frame, height=2, fg_color="#00b4d8")
        separator.pack(fill="x", padx=5, pady=(5, 0))

    def create_modern_field(self, parent, label, key, placeholder, bind_func=None):
        frame = ctk.CTkFrame(parent, fg_color="#293241", corner_radius=12)
        frame.pack(fill="x", pady=8)
        
        ctk.CTkLabel(frame, text=label, text_color="#98c1d9", 
                    font=ctk.CTkFont(weight="bold", size=13)).pack(anchor="w", padx=15, pady=(12, 6))
        
        entry = ctk.CTkEntry(frame, height=40, placeholder_text=placeholder,
                           font=ctk.CTkFont(size=13))
        entry.pack(fill="x", padx=15, pady=(0, 12))
        
        if bind_func:
            entry.bind('<KeyRelease>', bind_func)
        
        self.booking_inputs[key] = entry

    def open_calendar_picker(self):
        ModernCalendarDialog(self, self.update_date, self.selected_date)

    def update_date(self, date_str):
        try:
            date_obj = datetime.strptime(date_str, "%d-%m-%Y").date()
            self.selected_date = date_obj
            self.date_display.configure(state="normal")
            self.date_display.delete(0, 'end')
            self.date_display.insert(0, date_str)
            self.date_display.configure(state="readonly")
        except:
            pass

    def on_resort_change(self, choice):
        if choice == "--- Enter Custom Resort ---":
            self.custom_resort_entry.configure(state="normal")
            self.custom_resort_entry.focus()
        else:
            self.custom_resort_entry.delete(0, 'end')
            self.custom_resort_entry.configure(state="disabled")

    def get_resort_name(self):
        selected = self.resort_var.get()
        if selected == "--- Enter Custom Resort ---":
            custom = self.custom_resort_entry.get().strip()
            if not custom:
                raise ValueError("Please enter a custom resort name!")
            return custom
        return selected

    def check_phone_length(self, event):
        phone = self.booking_inputs["Phone Number"].get().strip()
        if len(phone) == 10 and phone.isdigit():
            self.auto_fill_guest_data()

    def auto_fill_guest_data(self):
        phone = self.booking_inputs["Phone Number"].get().strip()
        if len(phone) == 10:
            bookings = self.db.search_guest(phone)
            if bookings:
                last_booking = bookings[0]
                if messagebox.askyesno("Guest Found", 
                    f"Found existing guest: {last_booking[4]}\n\nAuto-fill guest information?"):
                    self.booking_inputs["Guest Name"].delete(0, 'end')
                    self.booking_inputs["Guest Name"].insert(0, last_booking[4])
                    self.booking_inputs["Aadhaar Card Number"].delete(0, 'end')
                    self.booking_inputs["Aadhaar Card Number"].insert(0, last_booking[6])
                    if len(last_booking) > 13 and last_booking[13]:
                        self.booking_inputs["Email"].delete(0, 'end')
                        self.booking_inputs["Email"].insert(0, last_booking[13])

    def toggle_activity_dropdown(self, choice=None):
        if self.interested_var.get() == "Yes":
            self.activity_combo.configure(state="normal")
            if self.activity_var.get() == "None":
                self.activity_var.set("Kayaking")
        else:
            self.activity_var.set("None")
            self.activity_combo.configure(state="disabled")

    def save_booking(self):
        try:
            date_iso = self.selected_date.isoformat()
            
            try:
                resort = self.get_resort_name()
            except ValueError as e:
                messagebox.showerror("Resort Error", str(e))
                return
            
            members_str = self.booking_inputs["Members"].get().strip()
            amount_str = self.booking_inputs["Amount Paid"].get().strip()
            profit_str = self.booking_inputs["Profit"].get().strip()
            advance_str = self.booking_inputs["Advance Paid"].get().strip()
            balance_str = self.booking_inputs["Balance Due"].get().strip()
            
            if not members_str or not amount_str or not profit_str:
                messagebox.showerror("Missing Information", 
                    "Please fill in Members, Amount, and Profit!")
                return
            
            try:
                members = int(members_str)
                amount = float(amount_str)
                profit = float(profit_str)
                advance = float(advance_str) if advance_str else 0
                balance = float(balance_str) if balance_str else 0
            except ValueError:
                messagebox.showerror("Invalid Input", 
                    "Please enter valid numbers for numeric fields!")
                return
            
            if members <= 0 or amount < 0 or profit < 0:
                messagebox.showerror("Invalid Values", "Please check your numeric values!")
                return
            
            guest = self.booking_inputs["Guest Name"].get().strip()
            phone = self.booking_inputs["Phone Number"].get().strip()
            aadhaar = self.booking_inputs["Aadhaar Card Number"].get().strip()
            email = self.booking_inputs["Email"].get().strip()
            
            if not guest or not phone or not aadhaar:
                messagebox.showerror("Missing Information", 
                    "Please fill all required guest information fields!")
                return
            
            phone = re.sub(r'\D', '', phone)
            aadhaar = re.sub(r'\D', '', aadhaar)
            
            if len(phone) != 10:
                messagebox.showerror("Invalid Phone", 
                    "Phone number must be exactly 10 digits!")
                return
            
            if len(aadhaar) != 12:
                messagebox.showerror("Invalid Aadhaar", 
                    "Aadhaar number must be exactly 12 digits!")
                return
            
            interested = self.interested_var.get()
            activity = self.activity_var.get() if interested == "Yes" else "None"
            notes = self.notes_textbox.get("1.0", "end-1c").strip()
            special_requests = self.special_requests_textbox.get("1.0", "end-1c").strip()
            status = self.status_var.get()
            payment_mode = self.payment_var.get()
            check_in = self.booking_inputs["Check-in Time"].get().strip()
            check_out = self.booking_inputs["Check-out Time"].get().strip()

            booking_data = (date_iso, members, amount, guest, phone, aadhaar, resort, 
                          interested, activity, profit, notes, email, status, payment_mode,
                          advance, balance, check_in, check_out, special_requests)
            
            booking_id = self.db.insert(booking_data)
            
            # Send notifications
            if (self.send_email_var.get() or self.send_sms_var.get()) and email:
                notification_data = {
                    'booking_id': booking_id,
                    'guest': guest,
                    'resort': resort,
                    'date': self.selected_date.strftime("%d-%m-%Y"),
                    'members': members,
                    'amount': amount,
                    'email': email,
                    'phone': phone
                }
                
                if self.send_email_var.get() and email and EMAIL_CONFIG['enabled']:
                    if NotificationService.send_booking_confirmation(notification_data):
                        self.db.log_activity("EMAIL_SENT", f"Confirmation sent to {email}", USER)
            
            messagebox.showinfo("✅ Success", 
                f"Booking #{booking_id} saved successfully!\n\n"
                f"Guest: {guest}\nResort: {resort}\nAmount: ₹{amount:,.2f}")
            
            # Clear form
            for e in self.booking_inputs.values():
                e.delete(0, 'end')
            self.resort_var.set(self.RESORT_LIST[0])
            self.custom_resort_entry.delete(0, 'end')
            self.custom_resort_entry.configure(state="disabled")
            self.interested_var.set("No")
            self.activity_var.set("None")
            self.activity_combo.configure(state="disabled")
            self.status_var.set("Confirmed")
            self.payment_var.set("Cash")
            self.notes_textbox.delete("1.0", "end")
            self.special_requests_textbox.delete("1.0", "end")
            self.selected_date = datetime.now().date()
            self.date_display.configure(state="normal")
            self.date_display.delete(0, 'end')
            self.date_display.insert(0, self.selected_date.strftime("%d-%m-%Y"))
            self.date_display.configure(state="readonly")
            
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred:\n{str(e)}")

    def create_history_ui(self):
        for widget in self.history_frame.winfo_children():
            widget.destroy()

        # Security Badge
        badge_frame = ctk.CTkFrame(self.history_frame, fg_color="transparent", height=40)
        badge_frame.pack(pady=(18, 5))
        badge_frame.pack_propagate(False)
        
        badge = ctk.CTkFrame(badge_frame, fg_color="#1a2332", corner_radius=20)
        badge.pack()
        ctk.CTkLabel(badge, text="🔒 Secure Area - PIN Protected", 
                    font=ctk.CTkFont(size=13, weight="bold"), 
                    text_color="#90e0ef").pack(padx=20, pady=8)

        # Header
        header = ctk.CTkFrame(self.history_frame, fg_color="#1a2332", height=75)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        ctk.CTkLabel(header, text="📚 Booking History", 
                    font=ctk.CTkFont(size=30, weight="bold"), 
                    text_color="#00d4ff").pack(pady=22)

        # Advanced Filters Card
        filter_card = ctk.CTkFrame(self.history_frame, fg_color="#1a2332", corner_radius=15)
        filter_card.pack(pady=15, padx=45, fill="x")

        ctk.CTkLabel(filter_card, text="🔍 Advanced Filters", 
                    font=ctk.CTkFont(size=17, weight="bold"), 
                    text_color="#00d4ff").pack(anchor="w", padx=20, pady=(18, 12))

        filter_frame = ctk.CTkFrame(filter_card, fg_color="transparent")
        filter_frame.pack(padx=20, pady=(0, 18))

        # Row 1 - Text Filters
        row1 = ctk.CTkFrame(filter_frame, fg_color="transparent")
        row1.pack(fill="x", pady=6)

        ctk.CTkLabel(row1, text="Guest/Phone:", text_color="#98c1d9",
                    font=ctk.CTkFont(size=12, weight="bold")).grid(row=0, column=0, padx=5, sticky="w")
        self.guest_filter = ctk.CTkEntry(row1, width=140, height=35, font=ctk.CTkFont(size=12))
        self.guest_filter.grid(row=0, column=1, padx=5)

        ctk.CTkLabel(row1, text="Resort:", text_color="#98c1d9",
                    font=ctk.CTkFont(size=12, weight="bold")).grid(row=0, column=2, padx=5, sticky="w")
        self.resort_filter = ctk.CTkEntry(row1, width=140, height=35, font=ctk.CTkFont(size=12))
        self.resort_filter.grid(row=0, column=3, padx=5)

        ctk.CTkLabel(row1, text="Status:", text_color="#98c1d9",
                    font=ctk.CTkFont(size=12, weight="bold")).grid(row=0, column=4, padx=5, sticky="w")
        self.status_filter_var = ctk.StringVar(value="All")
        self.status_filter_combo = ctk.CTkComboBox(row1, 
                    values=["All", "Confirmed", "Pending", "Cancelled", "Completed"], 
                    variable=self.status_filter_var, width=130, height=35,
                    font=ctk.CTkFont(size=12))
        self.status_filter_combo.grid(row=0, column=5, padx=5)
        
        ctk.CTkLabel(row1, text="Activity:", text_color="#98c1d9",
                    font=ctk.CTkFont(size=12, weight="bold")).grid(row=0, column=6, padx=5, sticky="w")
        self.activity_filter_var = ctk.StringVar(value="All")
        self.activity_filter_combo = ctk.CTkComboBox(row1, 
                    values=["All", "None", "Kayaking", "Zorbing", "Swimming", "Boating", 
                           "Zip Line", "Short Rafting", "Mid Rafting", "Long Rafting", 
                           "Go Karting", "All Activities"], 
                    variable=self.activity_filter_var, width=140, height=35,
                    font=ctk.CTkFont(size=12))
        self.activity_filter_combo.grid(row=0, column=7, padx=5)

        # Row 2 - Date Filters
        row2 = ctk.CTkFrame(filter_frame, fg_color="transparent")
        row2.pack(fill="x", pady=6)

        ctk.CTkLabel(row2, text="From:", text_color="#98c1d9",
                    font=ctk.CTkFont(size=12, weight="bold")).grid(row=0, column=0, padx=5, sticky="w")
        self.start_date_filter = DateEntry(row2, date_pattern="dd-mm-yyyy", 
                                          width=13, font=("Segoe UI", 10))
        self.start_date_filter.grid(row=0, column=1, padx=5)

        ctk.CTkLabel(row2, text="To:", text_color="#98c1d9",
                    font=ctk.CTkFont(size=12, weight="bold")).grid(row=0, column=2, padx=5, sticky="w")
        self.end_date_filter = DateEntry(row2, date_pattern="dd-mm-yyyy", 
                                        width=13, font=("Segoe UI", 10))
        self.end_date_filter.grid(row=0, column=3, padx=5)

        # Quick Filters
        quick_frame = ctk.CTkFrame(row2, fg_color="transparent")
        quick_frame.grid(row=0, column=4, columnspan=4, sticky="w", padx=10)
        
        ctk.CTkButton(quick_frame, text="Today", command=lambda: self.quick_filter(0), 
                     width=75, height=32, fg_color="#6a4c93", hover_color="#553978",
                     font=ctk.CTkFont(size=11, weight="bold")).pack(side="left", padx=3)
        ctk.CTkButton(quick_frame, text="Week", command=lambda: self.quick_filter(7), 
                     width=75, height=32, fg_color="#6a4c93", hover_color="#553978",
                     font=ctk.CTkFont(size=11, weight="bold")).pack(side="left", padx=3)
        ctk.CTkButton(quick_frame, text="Month", command=lambda: self.quick_filter(30), 
                     width=75, height=32, fg_color="#6a4c93", hover_color="#553978",
                     font=ctk.CTkFont(size=11, weight="bold")).pack(side="left", padx=3)

        # Action Buttons
        action_frame = ctk.CTkFrame(row2, fg_color="transparent")
        action_frame.grid(row=0, column=8, columnspan=2, sticky="e", padx=5)
        
        ctk.CTkButton(action_frame, text="Apply", command=self.apply_filters, 
                     width=80, height=32, fg_color="#00b4d8", hover_color="#0096c7",
                     font=ctk.CTkFont(size=12, weight="bold")).pack(side="left", padx=3)
        ctk.CTkButton(action_frame, text="Clear", command=self.clear_filters, 
                     width=80, height=32, fg_color="#e63946", hover_color="#d62828",
                     font=ctk.CTkFont(size=12, weight="bold")).pack(side="left", padx=3)

        # Table Card
        table_card = ctk.CTkFrame(self.history_frame, fg_color="#1a2332", corner_radius=15)
        table_card.pack(expand=True, fill="both", padx=45, pady=15)

        table_frame = ctk.CTkFrame(table_card, fg_color="transparent")
        table_frame.pack(expand=True, fill="both", padx=18, pady=18)

        # Enhanced table styling
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Treeview", 
                       background="#1a2332", 
                       foreground="#ffffff", 
                       fieldbackground="#1a2332", 
                       font=('Segoe UI', 10),
                       rowheight=28)
        style.configure("Treeview.Heading", 
                       background="#293241", 
                       foreground="#00d4ff", 
                       font=('Segoe UI', 11, 'bold'),
                       relief="flat")
        style.map('Treeview', 
                 background=[('selected', '#00b4d8')],
                 foreground=[('selected', 'white')])

        columns = ["ID", "Date", "Guest", "Phone", "Resort", "Members", "Amount", 
                  "Advance", "Balance", "Profit", "Payment", "Activity", "Status"]
        
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=14)
        
        column_widths = {
            "ID": 50, "Date": 90, "Guest": 120, "Phone": 95, "Resort": 130,
            "Members": 70, "Amount": 90, "Advance": 85, "Balance": 85,
            "Profit": 85, "Payment": 95, "Activity": 100, "Status": 85
        }
        
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=column_widths.get(col, 100), anchor='center')

        self.tree.bind('<Double-Button-1>', self.on_booking_double_click)
        self.tree.bind('<Button-3>', self.show_context_menu)

        scrollbar_y = ctk.CTkScrollbar(table_frame, orientation="vertical", command=self.tree.yview)
        scrollbar_x = ctk.CTkScrollbar(table_frame, orientation="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        self.tree.grid(row=0, column=0, sticky="nsew")
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        scrollbar_x.grid(row=1, column=0, sticky="ew")
        
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        self.context_menu = None

        # Action Buttons
        action_btn_frame = ctk.CTkFrame(self.history_frame, fg_color="transparent")
        action_btn_frame.pack(pady=12)

        ctk.CTkButton(action_btn_frame, text="✏️ Edit Selected", command=self.edit_selected_booking, 
                     width=140, height=38, fg_color="#00b4d8", hover_color="#0096c7",
                     font=ctk.CTkFont(size=13, weight="bold")).pack(side="left", padx=6)
        ctk.CTkButton(action_btn_frame, text="🗑️ Delete Selected", command=self.delete_selected_booking, 
                     width=140, height=38, fg_color="#e63946", hover_color="#d62828",
                     font=ctk.CTkFont(size=13, weight="bold")).pack(side="left", padx=6)
        ctk.CTkButton(action_btn_frame, text="📄 View Details", command=self.view_booking_details, 
                     width=140, height=38, fg_color="#6a4c93", hover_color="#553978",
                     font=ctk.CTkFont(size=13, weight="bold")).pack(side="left", padx=6)

        # Pagination
        pag_frame = ctk.CTkFrame(self.history_frame, fg_color="transparent")
        pag_frame.pack(pady=15)

        ctk.CTkButton(pag_frame, text="⬅ Prev", command=self.prev_page, 
                     width=110, height=38, fg_color="#00b4d8", hover_color="#0096c7",
                     font=ctk.CTkFont(size=13, weight="bold")).pack(side="left", padx=12)
        
        self.page_label = ctk.CTkLabel(pag_frame, text="Page 1 / 1", text_color="#00d4ff", 
                                       font=ctk.CTkFont(size=15, weight="bold"))
        self.page_label.pack(side="left", padx=25)
        
        ctk.CTkButton(pag_frame, text="Next ➡", command=self.next_page, 
                     width=110, height=38, fg_color="#00b4d8", hover_color="#0096c7",
                     font=ctk.CTkFont(size=13, weight="bold")).pack(side="left", padx=12)

    def quick_filter(self, days):
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=days)
        self.start_date_filter.set_date(start_date)
        self.end_date_filter.set_date(end_date)
        self.apply_filters()

    def show_context_menu(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            
            if self.context_menu:
                self.context_menu.destroy()
            
            self.context_menu = ctk.CTkToplevel(self)
            self.context_menu.wm_overrideredirect(True)
            self.context_menu.configure(fg_color="#1a2332", border_width=2, border_color="#00b4d8")
            
            x, y = event.x_root, event.y_root
            self.context_menu.geometry(f"+{x}+{y}")
            
            ctk.CTkButton(self.context_menu, text="✏️ Edit", command=self.edit_selected_booking, 
                         width=130, fg_color="#00b4d8", hover_color="#0096c7",
                         font=ctk.CTkFont(size=12, weight="bold")).pack(padx=5, pady=5)
            ctk.CTkButton(self.context_menu, text="📄 View Details", command=self.view_booking_details, 
                         width=130, fg_color="#6a4c93", hover_color="#553978",
                         font=ctk.CTkFont(size=12, weight="bold")).pack(padx=5, pady=5)
            ctk.CTkButton(self.context_menu, text="🗑️ Delete", command=self.delete_selected_booking, 
                         width=130, fg_color="#e63946", hover_color="#d62828",
                         font=ctk.CTkFont(size=12, weight="bold")).pack(padx=5, pady=5)
            
            self.after(100, lambda: self.context_menu.bind('<FocusOut>', lambda e: self.context_menu.destroy()))
            self.context_menu.focus()

    def view_booking_details(self):
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a booking to view!")
            return
        
        item = self.tree.item(selection[0])
        booking_id = item['values'][0]
        booking_data = self.db.get_booking_by_id(booking_id)
        
        if booking_data:
            self.show_booking_detail_dialog(booking_data)

    def show_booking_detail_dialog(self, booking):
        detail_window = ctk.CTkToplevel(self)
        detail_window.title(f"Booking Details - #{booking[0]}")
        detail_window.geometry("700x850")
        detail_window.configure(fg_color="#0a0e27")
        detail_window.transient(self)
        detail_window.grab_set()
        
        # Header
        header = ctk.CTkFrame(detail_window, fg_color="#1a2332", height=70)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        ctk.CTkLabel(header, text=f"📋 Booking Details - #{booking[0]}", 
                    font=ctk.CTkFont(size=24, weight="bold"), 
                    text_color="#00d4ff").pack(pady=20)
        
        scroll = ctk.CTkScrollableFrame(detail_window, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=30, pady=20)
        
        # Helper function to create detail rows
        def create_detail_row(parent, label, value, value_color="#ffffff"):
            row = ctk.CTkFrame(parent, fg_color="#293241", corner_radius=8)
            row.pack(fill="x", pady=4)
            
            ctk.CTkLabel(row, text=label, text_color="#98c1d9", 
                        font=ctk.CTkFont(size=13, weight="bold")).pack(side="left", padx=15, pady=10)
            ctk.CTkLabel(row, text=str(value), text_color=value_color, 
                        font=ctk.CTkFont(size=13)).pack(side="right", padx=15, pady=10)
        
        # Guest Information
        section_label = ctk.CTkLabel(scroll, text="👤 Guest Information", 
                                     font=ctk.CTkFont(size=18, weight="bold"), 
                                     text_color="#00d4ff")
        section_label.pack(anchor="w", pady=(10, 8))
        
        create_detail_row(scroll, "Guest Name:", booking[4])
        create_detail_row(scroll, "Phone:", booking[5])
        create_detail_row(scroll, "Email:", booking[13] if len(booking) > 13 and booking[13] else "N/A")
        create_detail_row(scroll, "Aadhaar:", booking[6])
        
        # Booking Information
        section_label = ctk.CTkLabel(scroll, text="📅 Booking Information", 
                                     font=ctk.CTkFont(size=18, weight="bold"), 
                                     text_color="#00d4ff")
        section_label.pack(anchor="w", pady=(20, 8))
        
        create_detail_row(scroll, "Resort:", booking[7])
        create_detail_row(scroll, "Date:", datetime.fromisoformat(booking[1]).strftime("%d-%m-%Y"))
        create_detail_row(scroll, "Members:", booking[2])
        create_detail_row(scroll, "Check-in Time:", booking[18] if len(booking) > 18 and booking[18] else "N/A")
        create_detail_row(scroll, "Check-out Time:", booking[19] if len(booking) > 19 and booking[19] else "N/A")
        create_detail_row(scroll, "Status:", booking[14] if len(booking) > 14 else "Confirmed", 
                         "#00d4ff" if (len(booking) > 14 and booking[14] == "Confirmed") else "#ffc107")
        
        # Financial Information
        section_label = ctk.CTkLabel(scroll, text="💰 Financial Information", 
                                     font=ctk.CTkFont(size=18, weight="bold"), 
                                     text_color="#00d4ff")
        section_label.pack(anchor="w", pady=(20, 8))
        
        create_detail_row(scroll, "Total Amount:", f"₹{booking[3]:,.2f}", "#00d4ff")
        create_detail_row(scroll, "Advance Paid:", f"₹{booking[16]:,.2f}" if len(booking) > 16 else "₹0.00", "#48cae4")
        create_detail_row(scroll, "Balance Due:", f"₹{booking[17]:,.2f}" if len(booking) > 17 else "₹0.00", "#ffc107")
        create_detail_row(scroll, "Profit:", f"₹{booking[10]:,.2f}", "#2e7d32")
        create_detail_row(scroll, "Payment Mode:", booking[15] if len(booking) > 15 else "Cash")
        
        # Activity Information
        section_label = ctk.CTkLabel(scroll, text="🌊 Activity Information", 
                                     font=ctk.CTkFont(size=18, weight="bold"), 
                                     text_color="#00d4ff")
        section_label.pack(anchor="w", pady=(20, 8))
        
        create_detail_row(scroll, "Interested:", booking[8])
        create_detail_row(scroll, "Activity:", booking[9])
        
        # Additional Information
        if (len(booking) > 12 and booking[12]) or (len(booking) > 20 and booking[20]):
            section_label = ctk.CTkLabel(scroll, text="📝 Additional Information", 
                                         font=ctk.CTkFont(size=18, weight="bold"), 
                                         text_color="#00d4ff")
            section_label.pack(anchor="w", pady=(20, 8))
            
            if len(booking) > 12 and booking[12]:
                notes_frame = ctk.CTkFrame(scroll, fg_color="#293241", corner_radius=8)
                notes_frame.pack(fill="x", pady=4)
                ctk.CTkLabel(notes_frame, text="Notes:", text_color="#98c1d9", 
                            font=ctk.CTkFont(size=13, weight="bold")).pack(anchor="w", padx=15, pady=(10, 5))
                ctk.CTkLabel(notes_frame, text=booking[12], text_color="#ffffff", 
                            font=ctk.CTkFont(size=12), wraplength=600, justify="left").pack(anchor="w", padx=15, pady=(0, 10))
            
            if len(booking) > 20 and booking[20]:
                requests_frame = ctk.CTkFrame(scroll, fg_color="#293241", corner_radius=8)
                requests_frame.pack(fill="x", pady=4)
                ctk.CTkLabel(requests_frame, text="Special Requests:", text_color="#98c1d9", 
                            font=ctk.CTkFont(size=13, weight="bold")).pack(anchor="w", padx=15, pady=(10, 5))
                ctk.CTkLabel(requests_frame, text=booking[20], text_color="#ffffff", 
                            font=ctk.CTkFont(size=12), wraplength=600, justify="left").pack(anchor="w", padx=15, pady=(0, 10))
        
        # Timestamp
        if len(booking) > 11 and booking[11]:
            section_label = ctk.CTkLabel(scroll, text="⏰ Timestamp", 
                                         font=ctk.CTkFont(size=18, weight="bold"), 
                                         text_color="#00d4ff")
            section_label.pack(anchor="w", pady=(20, 8))
            create_detail_row(scroll, "Created At:", booking[11])
        
        # Close Button
        ctk.CTkButton(scroll, text="Close", command=detail_window.destroy, 
                     width=200, height=45, fg_color="#e63946", hover_color="#d62828",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(pady=25)

    def on_booking_double_click(self, event):
        self.view_booking_details()

    def edit_selected_booking(self):
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a booking to edit!")
            return
        
        item = self.tree.item(selection[0])
        booking_id = item['values'][0]
        
        booking_data = self.db.get_booking_by_id(booking_id)
        if booking_data:
            EditBookingDialog(self, booking_data, lambda data: self.update_booking(booking_id, data))

    def update_booking(self, booking_id, data):
        try:
            booking_tuple = (
                data['date'], data['members'], data['amount'], data['guest'],
                data['phone'], data['aadhaar'], data['resort'], data['interested'],
                data['activity'], data['profit'], data['notes'], data['email'], 
                data['status'], data['payment_mode'], data['advance_paid'], 
                data['balance_due'], data['check_in_time'], data['check_out_time'], 
                data['special_requests']
            )
            self.db.update_booking(booking_id, booking_tuple)
            messagebox.showinfo("✅ Success", f"Booking #{booking_id} updated successfully!")
            self.load_history()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update booking:\n{str(e)}")

    def delete_selected_booking(self):
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a booking to delete!")
            return
        
        item = self.tree.item(selection[0])
        booking_id = item['values'][0]
        guest_name = item['values'][2]
        
        if messagebox.askyesno("Confirm Deletion", 
            f"Are you sure you want to delete booking #{booking_id}?\n\n"
            f"Guest: {guest_name}\n\nThis action cannot be undone!"):
            if self.db.delete_booking(booking_id):
                messagebox.showinfo("✅ Success", f"Booking #{booking_id} deleted successfully!")
                self.load_history()
            else:
                messagebox.showerror("Error", "Failed to delete booking!")

    def apply_filters(self):
        self.current_page = 1
        self.load_history()

    def clear_filters(self):
        self.guest_filter.delete(0, 'end')
        self.resort_filter.delete(0, 'end')
        self.status_filter_var.set("All")
        self.activity_filter_var.set("All")
        self.start_date_filter.set_date(datetime.now())
        self.end_date_filter.set_date(datetime.now())
        self.current_page = 1
        self.load_history()

    def load_history(self):
        self.tree.delete(*self.tree.get_children())
        
        try:
            guest_f = self.guest_filter.get().strip()
            resort_f = self.resort_filter.get().strip()
            start_f = self.start_date_filter.get_date().isoformat()
            end_f = self.end_date_filter.get_date().isoformat()
            status_f = "" if self.status_filter_var.get() == "All" else self.status_filter_var.get()
            activity_f = self.activity_filter_var.get()

            total_records = self.db.query_count(guest_f, resort_f, start_f, end_f, status_f, activity_f)
            self.total_pages = max(1, (total_records + self.page_size - 1) // self.page_size)
            self.current_page = min(max(1, self.current_page), self.total_pages)
            offset = (self.current_page - 1) * self.page_size

            rows = self.db.query_bookings(guest_f, resort_f, start_f, end_f, status_f, 
                                         activity_f, limit=self.page_size, offset=offset)
            
            for row in rows:
                display_row = [
                    row[0],  # ID
                    datetime.fromisoformat(row[1]).strftime("%d-%m-%Y"),  # Date
                    row[4],  # Guest
                    row[5],  # Phone
                    row[7],  # Resort
                    row[2],  # Members
                    f"₹{row[3]:,.2f}",  # Amount
                    f"₹{row[16]:,.2f}" if len(row) > 16 else "₹0.00",  # Advance
                    f"₹{row[17]:,.2f}" if len(row) > 17 else "₹0.00",  # Balance
                    f"₹{row[10]:,.2f}",  # Profit
                    row[15] if len(row) > 15 else "Cash",  # Payment Mode
                    row[9],  # Activity
                    row[14] if len(row) > 14 else "Confirmed"  # Status
                ]
                self.tree.insert('', 'end', values=display_row)
            
            self.page_label.configure(text=f"Page {self.current_page} / {self.total_pages} | Total: {total_records} bookings")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load history:\n{str(e)}")

    def prev_page(self):
        if self.current_page > 1:
            self.current_page -= 1
            self.load_history()

    def next_page(self):
        if self.current_page < self.total_pages:
            self.current_page += 1
            self.load_history()

    def create_summary_ui(self):
        for widget in self.summary_frame.winfo_children():
            widget.destroy()

        # Security Badge
        badge_frame = ctk.CTkFrame(self.summary_frame, fg_color="transparent", height=40)
        badge_frame.pack(pady=(18, 5))
        badge_frame.pack_propagate(False)
        
        badge = ctk.CTkFrame(badge_frame, fg_color="#1a2332", corner_radius=20)
        badge.pack()
        ctk.CTkLabel(badge, text="🔒 Secure Area - PIN Protected", 
                    font=ctk.CTkFont(size=13, weight="bold"), 
                    text_color="#90e0ef").pack(padx=20, pady=8)

        # Header
        header = ctk.CTkFrame(self.summary_frame, fg_color="#1a2332", height=75)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        ctk.CTkLabel(header, text="📊 Business Analytics", 
                    font=ctk.CTkFont(size=30, weight="bold"), 
                    text_color="#00d4ff").pack(pady=22)

        scroll = ctk.CTkScrollableFrame(self.summary_frame, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=45, pady=15)

        # Date Range Selector
        range_card = ctk.CTkFrame(scroll, fg_color="#1a2332", corner_radius=15)
        range_card.pack(fill="x", pady=12)
        
        range_frame = ctk.CTkFrame(range_card, fg_color="transparent")
        range_frame.pack(pady=18, padx=25)
        
        ctk.CTkLabel(range_frame, text="📅 Analysis Period:", text_color="#00d4ff", 
                    font=ctk.CTkFont(weight="bold", size=15)).pack(side="left", padx=12)
        
        self.analytics_range_var = ctk.StringVar(value="All Time")
        ctk.CTkComboBox(range_frame, 
                       values=["Today", "This Week", "This Month", "Last 30 Days", 
                              "Last 90 Days", "This Year", "All Time"], 
                       variable=self.analytics_range_var, width=170, height=38,
                       font=ctk.CTkFont(size=13),
                       command=lambda x: self.load_summary()).pack(side="left", padx=8)

        # Metrics Cards
        metrics_frame = ctk.CTkFrame(scroll, fg_color="transparent")
        metrics_frame.pack(pady=25)

        self.summary_labels = {}
        metrics = [
            ("📋", "Total Bookings", "#0077b6"),
            ("💰", "Total Revenue", "#00b4d8"),
            ("💵", "Total Profit", "#48cae4"),
            ("🌊", "Water Activities", "#90e0ef"),
            ("👥", "Total Members", "#023e8a"),
            ("💳", "Balance Due", "#ffc107")
        ]

        for idx, (icon, label, color) in enumerate(metrics):
            card = ctk.CTkFrame(metrics_frame, fg_color="#1a2332", corner_radius=15, 
                               width=260, height=130)
            card.grid(row=idx//3, column=idx%3, padx=15, pady=15)
            card.pack_propagate(False)

            ctk.CTkLabel(card, text=icon, font=ctk.CTkFont(size=32)).pack(pady=(18, 5))
            ctk.CTkLabel(card, text=label, font=ctk.CTkFont(size=13, weight="bold"), 
                        text_color="#98c1d9").pack()
            
            val_label = ctk.CTkLabel(card, text="...", font=ctk.CTkFont(size=26, weight="bold"), 
                                    text_color=color)
            val_label.pack(pady=(5, 18))
            self.summary_labels[label] = val_label

        # Charts Section
        charts_card = ctk.CTkFrame(scroll, fg_color="#1a2332", corner_radius=15)
        charts_card.pack(fill="both", expand=True, pady=25)

        ctk.CTkLabel(charts_card, text="📈 Revenue & Profit Trends (Last 30 Days)", 
                    font=ctk.CTkFont(size=19, weight="bold"), 
                    text_color="#00d4ff").pack(pady=(22, 12))

        self.chart_frame = ctk.CTkFrame(charts_card, fg_color="transparent")
        self.chart_frame.pack(fill="both", expand=True, padx=25, pady=(0, 22))

        # Activity Breakdown
        activity_card = ctk.CTkFrame(scroll, fg_color="#1a2332", corner_radius=15)
        activity_card.pack(fill="x", pady=12)

        ctk.CTkLabel(activity_card, text="🎯 Activity Performance", 
                    font=ctk.CTkFont(size=19, weight="bold"), 
                    text_color="#00d4ff").pack(anchor="w", padx=25, pady=(22, 12))

        self.activity_frame = ctk.CTkFrame(activity_card, fg_color="transparent")
        self.activity_frame.pack(fill="x", padx=25, pady=(0, 22))

        # Export Section
        export_card = ctk.CTkFrame(scroll, fg_color="#1a2332", corner_radius=15)
        export_card.pack(fill="x", pady=12)

        ctk.CTkLabel(export_card, text="📤 Export Data", 
                    font=ctk.CTkFont(size=21, weight="bold"), 
                    text_color="#00d4ff").pack(pady=(28, 12))
        ctk.CTkLabel(export_card, text="Generate comprehensive Excel report with analytics and charts", 
                    font=ctk.CTkFont(size=13), text_color="#98c1d9").pack(pady=(0, 18))
        
        export_info = ctk.CTkFrame(export_card, fg_color="#293241", corner_radius=10)
        export_info.pack(fill="x", padx=25, pady=(0, 18))
        
        ctk.CTkLabel(export_info, text="🔐 Excel files are password-protected for security", 
                    font=ctk.CTkFont(size=12), text_color="#90e0ef").pack(pady=12)
        
        ctk.CTkButton(export_card, text="📥 Generate Encrypted Excel Report", 
                     command=self.export_to_excel, width=280, height=48, 
                     font=ctk.CTkFont(size=15, weight="bold"), 
                     fg_color="#00b4d8", hover_color="#0096c7",
                     corner_radius=12).pack(pady=(0, 28))

    def load_summary(self):
        try:
            range_val = self.analytics_range_var.get()
            start_date = None
            end_date = datetime.now().date().isoformat()
            
            if range_val == "Today":
                start_date = end_date
            elif range_val == "This Week":
                start_date = (datetime.now().date() - timedelta(days=7)).isoformat()
            elif range_val == "This Month":
                start_date = (datetime.now().date() - timedelta(days=30)).isoformat()
            elif range_val == "Last 30 Days":
                start_date = (datetime.now().date() - timedelta(days=30)).isoformat()
            elif range_val == "Last 90 Days":
                start_date = (datetime.now().date() - timedelta(days=90)).isoformat()
            elif range_val == "This Year":
                start_date = f"{datetime.now().year}-01-01"
            
            sums = self.db.summary(start_date, end_date)
            self.summary_labels["Total Bookings"].configure(text=f"{sums['bookings']}")
            self.summary_labels["Total Revenue"].configure(text=f"₹{sums['revenue']:,.0f}")
            self.summary_labels["Total Profit"].configure(text=f"₹{sums['profit']:,.0f}")
            self.summary_labels["Water Activities"].configure(text=f"{sums['interested']}")
            self.summary_labels["Total Members"].configure(text=f"{sums['total_members']}")
            self.summary_labels["Balance Due"].configure(text=f"₹{sums['balance_due']:,.0f}")
            
            self.load_revenue_chart()
            self.load_activity_breakdown()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load summary:\n{str(e)}")

    def load_revenue_chart(self):
        for widget in self.chart_frame.winfo_children():
            widget.destroy()
        
        try:
            data = self.db.get_revenue_trend(30)
            
            if not data:
                ctk.CTkLabel(self.chart_frame, text="No data available for chart", 
                            text_color="#98c1d9", font=ctk.CTkFont(size=13)).pack(pady=25)
                return
            
            fig = Figure(figsize=(11, 5), facecolor='#1a2332')
            ax = fig.add_subplot(111)
            ax.set_facecolor('#1a2332')
            
            dates = [datetime.fromisoformat(d[0]) for d in data]
            revenues = [d[1] for d in data]
            profits = [d[2] for d in data]
            
            ax.plot(dates, revenues, marker='o', linewidth=2.5, color='#00b4d8', 
                   label='Revenue', markersize=6)
            ax.plot(dates, profits, marker='s', linewidth=2.5, color='#48cae4', 
                   label='Profit', markersize=6)
            
            ax.set_xlabel('Date', color='#98c1d9', fontsize=11, fontweight='bold')
            ax.set_ylabel('Amount (₹)', color='#98c1d9', fontsize=11, fontweight='bold')
            ax.tick_params(colors='#98c1d9', labelsize=9)
            ax.legend(facecolor='#293241', edgecolor='#00b4d8', labelcolor='#ffffff', 
                     fontsize=10, framealpha=0.9)
            ax.grid(True, alpha=0.2, color='#98c1d9', linestyle='--')
            
            for spine in ax.spines.values():
                spine.set_color('#98c1d9')
                spine.set_linewidth(1.5)
            
            fig.tight_layout()
            
            canvas = FigureCanvasTkAgg(fig, self.chart_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True)
            
        except Exception as e:
            ctk.CTkLabel(self.chart_frame, text=f"Chart error: {str(e)}", 
                        text_color="#e63946", font=ctk.CTkFont(size=12)).pack(pady=25)

    def load_activity_breakdown(self):
        for widget in self.activity_frame.winfo_children():
            widget.destroy()
        
        try:
            activities = self.db.get_activity_breakdown()
            
            if not activities:
                ctk.CTkLabel(self.activity_frame, text="No activity data available", 
                            text_color="#98c1d9", font=ctk.CTkFont(size=13)).pack(pady=12)
                return
            
            for activity, count, revenue, profit in activities[:6]:
                item_frame = ctk.CTkFrame(self.activity_frame, fg_color="#293241", corner_radius=10)
                item_frame.pack(fill="x", pady=6)
                
                ctk.CTkLabel(item_frame, text=activity, font=ctk.CTkFont(size=15, weight="bold"), 
                            text_color="#ffffff").pack(side="left", padx=22, pady=14)
                
                info_frame = ctk.CTkFrame(item_frame, fg_color="transparent")
                info_frame.pack(side="right", padx=22)
                
                ctk.CTkLabel(info_frame, text=f"{count} bookings", 
                            font=ctk.CTkFont(size=12), text_color="#90e0ef").pack(side="left", padx=12)
                ctk.CTkLabel(info_frame, text=f"₹{revenue:,.0f}", 
                            font=ctk.CTkFont(size=13, weight="bold"), text_color="#00d4ff").pack(side="left", padx=12)
                ctk.CTkLabel(info_frame, text=f"Profit: ₹{profit:,.0f}", 
                            font=ctk.CTkFont(size=12, weight="bold"), text_color="#2e7d32").pack(side="left", padx=12)
                
        except Exception as e:
            ctk.CTkLabel(self.activity_frame, text=f"Error: {str(e)}", 
                        text_color="#e63946", font=ctk.CTkFont(size=12)).pack(pady=12)

    def export_to_excel(self):
        try:
            guest_f = self.guest_filter.get().strip() if hasattr(self, 'guest_filter') else ""
            resort_f = self.resort_filter.get().strip() if hasattr(self, 'resort_filter') else ""
            start_f = self.start_date_filter.get_date().isoformat() if hasattr(self, 'start_date_filter') else None
            end_f = self.end_date_filter.get_date().isoformat() if hasattr(self, 'end_date_filter') else None
            status_f = "" if not hasattr(self, 'status_filter_var') else ("" if self.status_filter_var.get() == "All" else self.status_filter_var.get())
            activity_f = self.activity_filter_var.get() if hasattr(self, 'activity_filter_var') else "All"

            rows = self.db.get_all_bookings(guest_f, resort_f, start_f, end_f, status_f, activity_f)

            if not rows:
                messagebox.showwarning("No Data", "No bookings found to export!")
                return

            df = pd.DataFrame(rows, columns=["ID", "Date", "Members", "Amount", "Guest", "Phone", 
                                            "AADHAAR", "Resort", "Interested", "Activity", "Profit", 
                                            "Created", "Notes", "Email", "Status", "Payment Mode", 
                                            "Advance Paid", "Balance Due", "Check-in Time", 
                                            "Check-out Time", "Special Requests"])

            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"Safari_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not filepath:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Bookings"

            # Enhanced Styles
            title_fill = PatternFill(start_color="003D82", end_color="003D82", fill_type="solid")
            title_font = Font(name='Calibri', size=20, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
            border = Border(left=Side(style='thin', color='CCCCCC'), 
                          right=Side(style='thin', color='CCCCCC'), 
                          top=Side(style='thin', color='CCCCCC'), 
                          bottom=Side(style='thin', color='CCCCCC'))

            # Title
            ws.merge_cells('A1:O1')
            title = ws['A1']
            title.value = "🌴 Your Trip Our Place - Comprehensive Booking Report"
            title.font = title_font
            title.fill = title_fill
            title.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 40

            # Info
            ws.merge_cells('A2:O2')
            info = ws['A2']
            info.value = f"Generated: {datetime.now().strftime('%d-%m-%Y %I:%M %p')} | Total Records: {len(df)} | Status: CONFIDENTIAL"
            info.font = Font(name='Calibri', size=11, italic=True, color="666666")
            info.alignment = Alignment(horizontal="center")
            ws.row_dimensions[2].height = 20

            # Headers
            headers = ["ID", "Date", "Guest", "Phone", "Email", "Resort", "Members", "Amount", 
                      "Advance", "Balance", "Profit", "Payment", "Activity", "Status", "Notes"]
            for idx, col in enumerate(headers, 1):
                cell = ws.cell(row=4, column=idx, value=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            # Data
            for r_idx, row in enumerate(df.itertuples(index=False), 5):
                display_data = [
                    row[0],  # ID
                    row[1],  # Date
                    row[4],  # Guest
                    row[5],  # Phone
                    row[13] if len(row) > 13 else "",  # Email
                    row[7],  # Resort
                    row[2],  # Members
                    row[3],  # Amount
                    row[16] if len(row) > 16 else 0,  # Advance
                    row[17] if len(row) > 17 else 0,  # Balance
                    row[10],  # Profit
                    row[15] if len(row) > 15 else "Cash",  # Payment Mode
                    row[9],  # Activity
                    row[14] if len(row) > 14 else "Confirmed",  # Status
                    row[12] if len(row) > 12 else ""  # Notes
                ]
                
                for c_idx, val in enumerate(display_data, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = border
                    
                    # Color coding for status
                    if c_idx == 14:  # Status column
                        if val == "Confirmed":
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            cell.font = Font(color="006100", bold=True)
                        elif val == "Pending":
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                            cell.font = Font(color="9C6500", bold=True)
                        elif val == "Cancelled":
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.font = Font(color="9C0006", bold=True)
                    
                    if c_idx in [8, 9, 10, 11]:  # Financial columns
                        cell.number_format = '₹#,##0.00'
                        cell.font = Font(bold=True)

            # Auto-adjust column widths
            auto_adjust_columns(ws)
# Summary Sheet
            summary_ws = wb.create_sheet(title="Summary")
            summary_ws.sheet_properties.tabColor = "00B4D8"
            
            # Summary Title
            summary_ws.merge_cells('A1:D1')
            summary_title = summary_ws['A1']
            summary_title.value = "📊 BUSINESS SUMMARY"
            summary_title.font = Font(name='Calibri', size=18, bold=True, color="FFFFFF")
            summary_title.fill = header_fill
            summary_title.alignment = Alignment(horizontal="center", vertical="center")
            summary_ws.row_dimensions[1].height = 35

            sums = self.db.summary(start_f, end_f)
            summary_data = [
                ("Total Bookings:", sums['bookings']),
                ("Total Revenue:", f"₹{sums['revenue']:,.2f}"),
                ("Total Profit:", f"₹{sums['profit']:,.2f}"),
                ("Advance Collected:", f"₹{sums['advance_paid']:,.2f}"),
                ("Balance Due:", f"₹{sums['balance_due']:,.2f}"),
                ("Water Activities:", sums['interested']),
                ("Total Members:", sums['total_members'])
            ]

            for idx, (label, value) in enumerate(summary_data, 3):
                label_cell = summary_ws.cell(row=idx, column=2, value=label)
                label_cell.font = Font(name='Calibri', size=13, bold=True)
                label_cell.alignment = Alignment(horizontal="right")
                
                value_cell = summary_ws.cell(row=idx, column=3, value=value)
                value_cell.font = Font(name='Calibri', size=13, bold=True, color="0070C0")
                value_cell.alignment = Alignment(horizontal="left")

            # Resort Statistics
            resort_stats = self.db.get_resort_stats()
            if resort_stats:
                summary_ws.merge_cells(f'A{len(summary_data)+5}:D{len(summary_data)+5}')
                resort_title = summary_ws[f'A{len(summary_data)+5}']
                resort_title.value = "🏨 TOP RESORTS"
                resort_title.font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
                resort_title.fill = PatternFill(start_color="6A4C93", end_color="6A4C93", fill_type="solid")
                resort_title.alignment = Alignment(horizontal="center", vertical="center")
                
                resort_headers = ["Resort", "Bookings", "Revenue", "Profit"]
                start_row = len(summary_data) + 7
                for idx, header in enumerate(resort_headers, 1):
                    cell = summary_ws.cell(row=start_row, column=idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border
                
                for r_idx, (resort, count, revenue, profit) in enumerate(resort_stats[:10], start_row+1):
                    summary_ws.cell(row=r_idx, column=1, value=resort).border = border
                    summary_ws.cell(row=r_idx, column=2, value=count).border = border
                    summary_ws.cell(row=r_idx, column=3, value=f"₹{revenue:,.2f}").border = border
                    summary_ws.cell(row=r_idx, column=4, value=f"₹{profit:,.2f}").border = border

            # Protect sheets with password
            ws.protection.sheet = True
            ws.protection.password = EXCEL_PASSWORD
            ws.protection.enable()
            
            summary_ws.protection.sheet = True
            summary_ws.protection.password = EXCEL_PASSWORD
            summary_ws.protection.enable()

            # Save workbook
            _append_excel_footer(wb)
            wb.save(filepath)
            
            # Log export activity
            self.db.log_activity("EXPORT_EXCEL", f"Exported {len(df)} records to {os.path.basename(filepath)}", USER)
            
            messagebox.showinfo("✅ Success", 
                f"Excel report generated successfully!\n\n"
                f"Location: {filepath}\n\n"
                f"🔐 Password: {EXCEL_PASSWORD}\n\n"
                f"Total Records: {len(df)}")

        except Exception as e:
            messagebox.showerror("Error", f"Export failed:\n{str(e)}")

    def create_backup_ui(self):
        for widget in self.backup_frame.winfo_children():
            widget.destroy()

        # Header
        header = ctk.CTkFrame(self.backup_frame, fg_color="#1a2332", height=75)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        ctk.CTkLabel(header, text="💾 Database Backup & Recovery", 
                    font=ctk.CTkFont(size=30, weight="bold"), 
                    text_color="#00d4ff").pack(pady=22)

        scroll = ctk.CTkScrollableFrame(self.backup_frame, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=45, pady=15)

        # Backup Actions Card
        action_card = ctk.CTkFrame(scroll, fg_color="#1a2332", corner_radius=15)
        action_card.pack(fill="x", pady=12)

        ctk.CTkLabel(action_card, text="🛡️ Backup Actions", 
                    font=ctk.CTkFont(size=21, weight="bold"), 
                    text_color="#00d4ff").pack(pady=(28, 18))

        btn_frame = ctk.CTkFrame(action_card, fg_color="transparent")
        btn_frame.pack(pady=(0, 28))

        ctk.CTkButton(btn_frame, text="💾 Create Backup Now", command=self.create_backup, 
                     width=220, height=50, font=ctk.CTkFont(size=15, weight="bold"), 
                     fg_color="#00b4d8", hover_color="#0096c7",
                     corner_radius=12).pack(side="left", padx=12)
        ctk.CTkButton(btn_frame, text="📂 Open Backup Folder", command=self.open_backup_folder, 
                     width=220, height=50, font=ctk.CTkFont(size=15, weight="bold"), 
                     fg_color="#6a4c93", hover_color="#553978",
                     corner_radius=12).pack(side="left", padx=12)

        # Info Card
        info_card = ctk.CTkFrame(scroll, fg_color="#293241", corner_radius=12)
        info_card.pack(fill="x", pady=12)
        
        ctk.CTkLabel(info_card, text="ℹ️ Backup Information", 
                    font=ctk.CTkFont(size=15, weight="bold"), 
                    text_color="#90e0ef").pack(anchor="w", padx=20, pady=(15, 10))
        
        info_text = """
        • Backups are automatically created on application startup
        • Only the last 30 backups are retained
        • Backup files are stored in the 'backups' folder
        • Use restore feature to recover from a previous backup
        """
        
        ctk.CTkLabel(info_card, text=info_text, 
                    font=ctk.CTkFont(size=12), 
                    text_color="#98c1d9", justify="left").pack(anchor="w", padx=20, pady=(0, 15))

        # Backup List Card
        list_card = ctk.CTkFrame(scroll, fg_color="#1a2332", corner_radius=15)
        list_card.pack(fill="both", expand=True, pady=12)

        ctk.CTkLabel(list_card, text="📋 Available Backups", 
                    font=ctk.CTkFont(size=19, weight="bold"), 
                    text_color="#00d4ff").pack(pady=(22, 12))

        table_frame = ctk.CTkFrame(list_card, fg_color="transparent")
        table_frame.pack(fill="both", expand=True, padx=22, pady=(0, 22))

        columns = ["Filename", "Date Created", "Size (KB)", "Actions"]
        
        self.backup_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=12)
        
        column_widths = {"Filename": 350, "Date Created": 180, "Size (KB)": 120, "Actions": 100}
        
        for col in columns:
            self.backup_tree.heading(col, text=col)
            self.backup_tree.column(col, width=column_widths.get(col, 150), anchor='center')

        scrollbar = ctk.CTkScrollbar(table_frame, orientation="vertical", command=self.backup_tree.yview)
        self.backup_tree.configure(yscrollcommand=scrollbar.set)
        
        self.backup_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Restore Button
        restore_frame = ctk.CTkFrame(list_card, fg_color="transparent")
        restore_frame.pack(pady=(12, 22))
        
        ctk.CTkButton(restore_frame, text="♻️ Restore Selected Backup", command=self.restore_backup, 
                     width=270, height=50, font=ctk.CTkFont(size=15, weight="bold"), 
                     fg_color="#e63946", hover_color="#d62828",
                     corner_radius=12).pack(side="left", padx=8)
        ctk.CTkButton(restore_frame, text="🗑️ Delete Selected", command=self.delete_backup, 
                     width=180, height=50, font=ctk.CTkFont(size=15, weight="bold"), 
                     fg_color="#6c757d", hover_color="#5a6268",
                     corner_radius=12).pack(side="left", padx=8)

    def load_backup_list(self):
        self.backup_tree.delete(*self.backup_tree.get_children())
        
        try:
            if not os.path.exists(BACKUP_DIR):
                return
            
            backups = sorted([f for f in os.listdir(BACKUP_DIR) 
                            if f.startswith('backup_') and f.endswith('.db')], reverse=True)
            
            for backup_file in backups:
                file_path = os.path.join(BACKUP_DIR, backup_file)
                file_size = os.path.getsize(file_path) / 1024  # KB
                
                try:
                    date_str = backup_file.replace('backup_', '').replace('.db', '')
                    date_obj = datetime.strptime(date_str, '%Y%m%d_%H%M%S')
                    formatted_date = date_obj.strftime('%d-%m-%Y %I:%M %p')
                except:
                    formatted_date = "Unknown"
                
                self.backup_tree.insert('', 'end', values=[backup_file, formatted_date, f"{file_size:.2f}", ""])
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load backup list:\n{str(e)}")

    def create_backup(self):
        try:
            backup_file = self.db.backup_database()
            self.db.log_activity("BACKUP_CREATED", f"Manual backup: {os.path.basename(backup_file)}", USER)
            messagebox.showinfo("✅ Success", 
                f"Backup created successfully!\n\n{os.path.basename(backup_file)}")
            self.load_backup_list()
        except Exception as e:
            messagebox.showerror("Error", f"Backup failed:\n{str(e)}")

    def open_backup_folder(self):
        try:
            if not os.path.exists(BACKUP_DIR):
                os.makedirs(BACKUP_DIR)
            
            if os.name == 'nt':  # Windows
                os.startfile(BACKUP_DIR)
            elif os.name == 'posix':  # macOS and Linux
                import subprocess
                subprocess.Popen(['xdg-open', BACKUP_DIR])
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open backup folder:\n{str(e)}")

    def delete_backup(self):
        selection = self.backup_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a backup to delete!")
            return
        
        item = self.backup_tree.item(selection[0])
        backup_filename = item['values'][0]
        backup_path = os.path.join(BACKUP_DIR, backup_filename)
        
        if messagebox.askyesno("Confirm Deletion", 
            f"Are you sure you want to delete this backup?\n\n{backup_filename}\n\nThis action cannot be undone!"):
            try:
                os.remove(backup_path)
                self.db.log_activity("BACKUP_DELETED", f"Deleted backup: {backup_filename}", USER)
                messagebox.showinfo("✅ Success", "Backup deleted successfully!")
                self.load_backup_list()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete backup:\n{str(e)}")

    def restore_backup(self):
        selection = self.backup_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a backup to restore!")
            return
        
        item = self.backup_tree.item(selection[0])
        backup_filename = item['values'][0]
        backup_path = os.path.join(BACKUP_DIR, backup_filename)
        
        if messagebox.askyesno("⚠️ Confirm Restore", 
            f"Are you sure you want to restore from:\n\n{backup_filename}\n\n"
            f"This will replace your current database.\n"
            f"All recent changes will be lost!\n\n"
            f"This action cannot be undone!"):
            
            try:
                self.db.restore_database(backup_path)
                self.db.log_activity("DATABASE_RESTORED", f"Restored from: {backup_filename}", USER)
                messagebox.showinfo("✅ Success", 
                    "Database restored successfully!\n\nPlease restart the application.")
                self.on_closing()
            except Exception as e:
                messagebox.showerror("Error", f"Restore failed:\n{str(e)}")

    def on_closing(self):
        if messagebox.askyesno("Exit", "Are you sure you want to exit?"):
            try:
                self.db.backup_database()  # Final backup before closing
                self.db.log_activity("APPLICATION_CLOSED", "User logged out", USER)
            except:
                pass
            self.db.close()
            self.destroy()


if __name__ == "__main__":
    app = BookingApp()
    app.mainloop()


def show_splash_and_start():
    import threading, time
    try:
        root = ctk.CTk()
        root.geometry('900x480')
        root.overrideredirect(True)
        root.configure(fg_color='#0b0c10')
        frame = ctk.CTkFrame(root, fg_color='#0b0c10')
        frame.pack(fill='both', expand=True)
        ctk.CTkLabel(frame, text='🦁 Safari Booking System', font=ctk.CTkFont(size=28, weight='bold'), text_color='#E0E0E0').pack(expand=True)
        ctk.CTkLabel(frame, text='Developed by Mohammed Ismail', font=ctk.CTkFont(size=12, slant='italic'), text_color='#B0BEC5').pack(pady=(0,30))
        root.update()
        def _start():
            time.sleep(2.5)
            try:
                root.destroy()
            except:
                pass
            app = BookingApp()
            app.mainloop()
        threading.Thread(target=_start, daemon=True).start()
        root.mainloop()
    except Exception:
        app = BookingApp()
        app.mainloop()

